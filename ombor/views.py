import datetime
import io
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from django.core.exceptions import PermissionDenied

from projects.models import Firma, Project, WorkSection
from projects.roles import is_pto, is_snab, user_firma, visible_firmas, visible_projects
from . import services


def _tahrir_mumkin(user):
    """Ombor hujjatlarini o'zgartirish huquqi (prorab FAQAT ko'radi)."""
    return bool(user.is_superuser or is_pto(user) or is_snab(user))


def _pto_kerak(user):
    """Ombor hujjatlarini PTO, SNABJENIYE yoki admin o'zgartiradi."""
    if not _tahrir_mumkin(user):
        raise PermissionDenied("Ombor hujjatlarini faqat PTO, snabjeniye yoki admin o'zgartiradi.")


def _visible_warehouses(user):
    """Foydalanuvchi ko'ra oladigan omborlar — ko'rinadigan obyektlariga
    tegishli omborlar (visible_projects bilan BIR XIL doira).
    Admin — hammasi. Izolyatsiya shu yerdan boshqariladi."""
    qs = Warehouse.objects.all()
    if user.is_superuser:
        return qs
    return qs.filter(project__in=visible_projects(user))


def _warehouse_yoki_403(user, warehouse):
    """Ombor foydalanuvchining ko'rish doirasiga kirmasa 403."""
    if user.is_superuser:
        return warehouse
    ok = bool(warehouse is not None
              and _visible_warehouses(user).filter(pk=warehouse.pk).exists())
    if not ok:
        raise PermissionDenied("Bu ombor sizga biriktirilgan obyektlarga tegishli emas.")
    return warehouse
from .models import (
    Issue, IssueItem, Material, Receipt, ReceiptImage, ReceiptItem, StockBalance,
    StockMovement, Supplier, Warehouse,
)


def _material_ol(qiymat, unit=""):
    """Material ID yoki NOM bilan berilishi mumkin. Nom bo'lsa — topamiz yoki YARATAMIZ.
    (Prixodda qo'lda yangi material kiritish uchun.)"""
    qiymat = (qiymat or "").strip()
    if not qiymat:
        return None
    if qiymat.isdigit():
        return Material.objects.filter(pk=int(qiymat)).first()
    m = Material.objects.filter(name__iexact=qiymat).first()
    if m:
        return m
    # Yangi material — noyob kod generatsiya qilamiz
    n = Material.objects.count() + 1
    while Material.objects.filter(code=f"M-{n:04d}").exists():
        n += 1
    return Material.objects.create(
        code=f"M-{n:04d}", name=qiymat, unit=(unit or "dona").strip()[:32],
    )


def _supplier_ol(qiymat):
    """Yetkazib beruvchi ID yoki NOM bilan berilishi mumkin.
    Nom bo'lsa — topamiz yoki YARATAMIZ (prixodda qo'lda yozish uchun)."""
    qiymat = (qiymat or "").strip()
    if not qiymat:
        return None
    if qiymat.isdigit():
        return Supplier.objects.filter(pk=int(qiymat)).first()
    s = Supplier.objects.filter(name__iexact=qiymat).first()
    if s:
        return s
    return Supplier.objects.create(name=qiymat[:255])


def _money(v):
    return f"{v:,.0f}".replace(",", " ")


def _dec(v):
    if v is None:
        return None
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _qty(v):
    s = f"{Decimal(v):,.3f}".replace(",", " ")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def _ombor_data(firma_id, project_id, warehouse_id, user=None):
    """Filtrlangan kirim/chiqim/qoldiq ma'lumoti + pul yig'indilari.
    user berilsa — faqat o'sha foydalanuvchi firmasidagi omborlar (admin=hammasi)."""
    moves = StockMovement.objects.all()
    balances = StockBalance.objects.all()
    # Izolyatsiya (majburiy — GET filtridan qat'i nazar): faqat ko'rinadigan omborlar
    if user is not None and not user.is_superuser:
        _wh = _visible_warehouses(user)
        moves = moves.filter(warehouse__in=_wh)
        balances = balances.filter(warehouse__in=_wh)
    if firma_id:
        moves = moves.filter(warehouse__project__firma_id=firma_id)
        balances = balances.filter(warehouse__project__firma_id=firma_id)
    if warehouse_id:
        moves = moves.filter(warehouse_id=warehouse_id)
        balances = balances.filter(warehouse_id=warehouse_id)
    if project_id:
        moves = moves.filter(warehouse__project_id=project_id)
        balances = balances.filter(warehouse__project_id=project_id)

    bal = {(b.warehouse_id, b.material_id): b for b in balances}
    warehouses = {w.id: w for w in Warehouse.objects.select_related("project", "project__firma")}
    materials = {m.id: m for m in Material.objects.all()}

    agg = moves.values("warehouse_id", "material_id").annotate(
        kirim=Sum("quantity", filter=Q(direction=StockMovement.IN)),
        chiqim=Sum("quantity", filter=Q(direction=StockMovement.OUT)),
    )

    rows = []
    jami_qoldiq_val = Decimal("0")
    tugagan = 0
    for a in agg:
        wid, mid = a["warehouse_id"], a["material_id"]
        kirim = a["kirim"] or Decimal("0")
        chiqim = -(a["chiqim"] or Decimal("0"))
        b = bal.get((wid, mid))
        qoldiq_qty = b.quantity if b else (kirim - chiqim)
        qoldiq_val = b.total_value if b else Decimal("0")
        w = warehouses.get(wid)
        m = materials.get(mid)
        jami_qoldiq_val += qoldiq_val
        kam = qoldiq_qty <= 0
        if kam:
            tugagan += 1
        firma_nom = "—"
        if w and w.project_id and w.project.firma_id:
            firma_nom = w.project.firma.name
        avg = b.avg_cost if b else Decimal("0")
        rows.append({
            "firma": firma_nom,
            "ombor": w.kod_nom if w else "—",
            "loyiha": (w.project.code if w and w.project_id else "markaziy"),
            "material": m.name if m else "—",
            "birlik": m.unit if m else "",
            "kirim": _qty(kirim), "chiqim": _qty(chiqim), "qoldiq": _qty(qoldiq_qty),
            "kam": kam,
            "avg_str": _money(avg), "qoldiq_val_str": _money(qoldiq_val),
            # xom (eksport uchun)
            "kirim_n": float(kirim), "chiqim_n": float(chiqim), "qoldiq_n": float(qoldiq_qty),
            "avg_n": float(avg), "qval_n": float(qoldiq_val),
        })
    rows.sort(key=lambda r: (r["ombor"], r["material"]))

    kv = moves.filter(direction=StockMovement.IN).aggregate(s=Sum("total_cost"))["s"] or Decimal("0")
    cv = moves.filter(direction=StockMovement.OUT).aggregate(s=Sum("total_cost"))["s"] or Decimal("0")
    return {
        "rows": rows,
        "jami_qoldiq_val": jami_qoldiq_val,
        "kirim_val": kv,
        "chiqim_val": -cv,
        "tugagan": tugagan,
    }


@login_required
def ombor(request):
    """Material bo'yicha kirim · chiqim · qoldiq + pul KPI + kam qoldiq ogohlantirish."""
    firma_id = request.GET.get("firma") or ""
    project_id = request.GET.get("project") or ""
    warehouse_id = request.GET.get("warehouse") or ""
    d = _ombor_data(firma_id, project_id, warehouse_id, user=request.user)

    q = request.GET.urlencode()
    kontekst = {
        "rows": d["rows"],
        "firmalar": visible_firmas(request.user).order_by("name"),
        "projects": visible_projects(request.user).order_by("code"),
        "warehouses": _visible_warehouses(request.user).select_related("project").order_by("name"),
        "sel_firma": firma_id,
        "sel_project": project_id,
        "sel_warehouse": warehouse_id,
        "jami_qoldiq_val_str": _money(d["jami_qoldiq_val"]),
        "kirim_val_str": _money(d["kirim_val"]),
        "chiqim_val_str": _money(d["chiqim_val"]),
        "tugagan": d["tugagan"],
        "soni": len(d["rows"]),
        "query": q,
    }
    return render(request, "ombor/ombor.html", kontekst)


def _ombor_wb(firma="", project="", warehouse="", user=None):
    """Ombor jadvalini (ixtiyoriy filtr bilan) openpyxl Workbook qaytaradi."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    d = _ombor_data(firma or "", project or "", warehouse or "", user=user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ombor"
    headers = ["Firma", "Ombor", "Loyiha", "Material", "Birlik", "Kirim", "Chiqim", "Qoldiq", "O'rtacha narx", "Qoldiq summasi"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
    for r in d["rows"]:
        ws.append([r["firma"], r["ombor"], r["loyiha"], r["material"], r["birlik"],
                   r["kirim_n"], r["chiqim_n"], r["qoldiq_n"], r["avg_n"], r["qval_n"]])
    for i, w in enumerate([26, 14, 12, 28, 8, 12, 12, 12, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return wb


@login_required
def ombor_export(request):
    """Ombor jadvalini (filtr bilan) Excel'ga eksport."""
    wb = _ombor_wb(request.GET.get("firma") or "", request.GET.get("project") or "",
                   request.GET.get("warehouse") or "", user=request.user)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="ombor.xlsx"'
    return resp


# ===================== PRIXOD (kirim) =====================
@login_required
def prixod_list(request):
    firma_id = request.GET.get("firma") or ""
    project_id = request.GET.get("project") or ""
    receipts = (Receipt.objects.prefetch_related("images").select_related(
        "warehouse", "supplier", "warehouse__project", "warehouse__project__firma")
        .prefetch_related("items"))
    if not request.user.is_superuser:
        receipts = receipts.filter(warehouse__in=_visible_warehouses(request.user))
    if firma_id:
        receipts = receipts.filter(warehouse__project__firma_id=firma_id)
    if project_id:
        receipts = receipts.filter(warehouse__project_id=project_id)
    receipts = receipts.order_by("-date", "-id")
    projects = visible_projects(request.user).order_by("code")
    if firma_id:
        projects = projects.filter(firma_id=firma_id)
    rows = []
    for r in receipts:
        firma_nom = "—"
        if r.warehouse.project_id and r.warehouse.project.firma_id:
            firma_nom = r.warehouse.project.firma.name
        sup = "—"
        if r.supplier_id:
            sup = r.supplier.name
            if r.supplier.inn:
                sup += f" · STIR {r.supplier.inn}"
        rasmlar = []
        if r.image:
            rasmlar.append({"url": r.image.url, "nak": False, "summa_str": ""})
        for im in r.images.all():
            rasmlar.append({
                "url": im.image.url, "nak": im.turi == "nakladnoy",
                "summa_str": _money(im.summa) if im.summa else "",
            })
        rows.append({
            "obj": r, "total_str": _money(r.total),
            "supplier": sup,
            "firma": firma_nom,
            "rasmlar": rasmlar,
        })
    return render(request, "ombor/prixod.html", {
        "can_edit": _tahrir_mumkin(request.user),
        "rows": rows,
        "firmalar": visible_firmas(request.user).order_by("name"),
        "obyektlar": projects,
        "sel_firma": firma_id,
        "sel_project": project_id,
        "warehouses": _visible_warehouses(request.user).select_related("project").order_by("name"),
        "suppliers": Supplier.objects.order_by("name"),
        "materials": Material.objects.order_by("name"),
        "today": datetime.date.today().isoformat(),
    })


@login_required
def prixod_add(request):
    _pto_kerak(request.user)
    if request.method == "POST":
        wh = request.POST.get("warehouse")
        if not wh:
            messages.error(request, "Ombor tanlang.")
            return redirect("prixod_list")
        # Firma izolyatsiyasi: tanlangan ombor foydalanuvchi firmasiga tegishli bo'lsin
        _warehouse_yoki_403(request.user, get_object_or_404(Warehouse, pk=wh))
        try:
            date = datetime.date.fromisoformat(request.POST.get("date", ""))
        except ValueError:
            messages.error(request, "Sanani to'g'ri kiriting.")
            return redirect("prixod_list")
        # Yetkazib beruvchi: ro'yxatdan tanlangan NOM yoki qo'lda yozilgan YANGI nom
        yetkazuvchi = _supplier_ol(request.POST.get("supplier"))
        r = Receipt.objects.create(
            warehouse_id=wh, supplier=yetkazuvchi,
            date=date, doc_number=(request.POST.get("doc_number") or "").strip(),
            note=(request.POST.get("note") or "").strip(),
        )
        # Mahsulot rasmlari — 5 tagacha
        for f in request.FILES.getlist("images")[:5]:
            ReceiptImage.objects.create(receipt=r, turi="product", image=f)
        # Nakladnoy rasmlari — har biri ITOGO summasi bilan (ketma-ket qatorlar)
        for i in range(15):
            nf = request.FILES.get(f"nak_image_{i}")
            if nf is None:
                continue
            ReceiptImage.objects.create(
                receipt=r, turi="nakladnoy", image=nf,
                summa=_dec(request.POST.get(f"nak_summa_{i}")),
            )
        mats = request.POST.getlist("material")
        munits = request.POST.getlist("munit")
        qtys = request.POST.getlist("quantity")
        prices = request.POST.getlist("unit_price")
        n = 0
        for i in range(len(mats)):
            if not (mats[i] or "").strip():
                continue
            q = _dec(qtys[i] if i < len(qtys) else None)
            pr = _dec(prices[i] if i < len(prices) else None)
            if q is None or q <= 0 or pr is None or pr < 0:
                continue
            # Material ID yoki qo'lda yozilgan NOM — topamiz yoki yaratamiz
            mat = _material_ol(mats[i], munits[i] if i < len(munits) else "")
            if mat is None:
                continue
            ReceiptItem.objects.create(receipt=r, material=mat, quantity=q, unit_price=pr)
            n += 1
        if n == 0:
            r.delete()
            messages.error(request, "Kamida bitta to'liq material qatori kiriting.")
        else:
            messages.success(request, f"Prixod #{r.id} yaratildi ({n} qator). «Qayd qilish»ni bosing — omborga tushadi.")
    return redirect("prixod_list")


@login_required
def prixod_action(request, pk):
    _pto_kerak(request.user)
    r = get_object_or_404(Receipt, pk=pk)
    _warehouse_yoki_403(request.user, r.warehouse)
    if request.method == "POST":
        a = request.POST.get("action")
        try:
            if a == "post":
                services.post_receipt(r)
                messages.success(request, f"Prixod #{r.id} qayd qilindi — ombor to'ldirildi.")
                # Arxiv botga (admin + obyekt snabjeniyechilari) — xato bo'lsa jim
                try:
                    from .telegram_arxiv import arxiv_prixod
                    arxiv_prixod(r)
                except Exception:
                    pass
            elif a == "unpost":
                services.unpost_receipt(r)
                messages.info(request, f"Prixod #{r.id} qaydi bekor qilindi.")
            elif a == "delete":
                if r.is_posted:
                    messages.error(request, "Avval qaydni bekor qiling.")
                else:
                    r.delete()
                    messages.info(request, "Prixod o'chirildi.")
        except ValidationError as e:
            messages.error(request, getattr(e, "message", str(e)))
    return redirect("prixod_list")


# ===================== RASXOD (chiqim) =====================
@login_required
def rasxod_list(request):
    firma_id = request.GET.get("firma") or ""
    project_id = request.GET.get("project") or ""
    issues = (Issue.objects.select_related(
        "warehouse", "work_section", "work_section__project",
        "warehouse__project", "warehouse__project__firma").prefetch_related("items"))
    if not request.user.is_superuser:
        issues = issues.filter(warehouse__in=_visible_warehouses(request.user))
    if firma_id:
        issues = issues.filter(warehouse__project__firma_id=firma_id)
    if project_id:
        issues = issues.filter(warehouse__project_id=project_id)
    issues = issues.order_by("-date", "-id")
    projects = visible_projects(request.user).order_by("code")
    if firma_id:
        projects = projects.filter(firma_id=firma_id)
    rows = []
    for x in issues:
        firma_nom = "—"
        if x.warehouse.project_id and x.warehouse.project.firma_id:
            firma_nom = x.warehouse.project.firma.name
        rows.append({
            "obj": x, "total_str": _money(x.total),
            "bolim": (x.work_section.name if x.work_section_id else "—"),
            "firma": firma_nom,
        })
    return render(request, "ombor/rasxod.html", {
        "can_edit": _tahrir_mumkin(request.user),
        "rows": rows,
        "firmalar": visible_firmas(request.user).order_by("name"),
        "obyektlar": projects,
        "sel_firma": firma_id,
        "sel_project": project_id,
        "warehouses": _visible_warehouses(request.user).select_related("project").order_by("name"),
        "sections": WorkSection.objects.filter(
            project__in=visible_projects(request.user)).select_related("project").order_by("project__code", "name"),
        "materials": Material.objects.order_by("name"),
        "today": datetime.date.today().isoformat(),
    })


@login_required
def rasxod_add(request):
    _pto_kerak(request.user)
    if request.method == "POST":
        wh = request.POST.get("warehouse")
        if not wh:
            messages.error(request, "Ombor tanlang.")
            return redirect("rasxod_list")
        # Firma izolyatsiyasi: ombor foydalanuvchi firmasiga tegishli bo'lsin
        _warehouse_yoki_403(request.user, get_object_or_404(Warehouse, pk=wh))
        # Ish bo'limi (bo'lsa) ham o'z firmasi obyektiga tegishli bo'lsin
        ws_id = request.POST.get("work_section") or None
        if ws_id and not request.user.is_superuser:
            ws = get_object_or_404(WorkSection, pk=ws_id)
            if ws.project_id not in set(visible_projects(request.user).values_list("id", flat=True)):
                raise PermissionDenied("Bu ish bo'limi sizning firmangizga tegishli emas.")
        try:
            date = datetime.date.fromisoformat(request.POST.get("date", ""))
        except ValueError:
            messages.error(request, "Sanani to'g'ri kiriting.")
            return redirect("rasxod_list")
        x = Issue.objects.create(
            warehouse_id=wh, work_section_id=ws_id,
            recipient=(request.POST.get("recipient") or "").strip(),
            date=date, doc_number=(request.POST.get("doc_number") or "").strip(),
            note=(request.POST.get("note") or "").strip(),
        )
        mats = request.POST.getlist("material")
        qtys = request.POST.getlist("quantity")
        n = 0
        for i in range(len(mats)):
            if not mats[i]:
                continue
            q = _dec(qtys[i] if i < len(qtys) else None)
            if q is None or q <= 0:
                continue
            # Material ID tekshirilib olinadi (noto'g'ri qiymat 500 xato bermasin)
            mat = Material.objects.filter(pk=mats[i]).first() if str(mats[i]).isdigit() else None
            if mat is None:
                continue
            IssueItem.objects.create(issue=x, material=mat, quantity=q)
            n += 1
        if n == 0:
            x.delete()
            messages.error(request, "Kamida bitta material qatori kiriting.")
        else:
            messages.success(request, f"Rasxod #{x.id} yaratildi ({n} qator). «Qayd qilish»ni bosing — ombordan chiqadi.")
    return redirect("rasxod_list")


@login_required
def rasxod_action(request, pk):
    _pto_kerak(request.user)
    x = get_object_or_404(Issue, pk=pk)
    _warehouse_yoki_403(request.user, x.warehouse)
    if request.method == "POST":
        a = request.POST.get("action")
        try:
            if a == "post":
                services.post_issue(x)
                messages.success(request, f"Rasxod #{x.id} qayd qilindi — ombordan chiqarildi.")
                # Arxiv botga (admin + obyekt snabjeniyechilari) — xato bo'lsa jim
                try:
                    from .telegram_arxiv import arxiv_rasxod
                    arxiv_rasxod(x)
                except Exception:
                    pass
            elif a == "unpost":
                services.unpost_issue(x)
                messages.info(request, f"Rasxod #{x.id} qaydi bekor qilindi.")
            elif a == "delete":
                if x.is_posted:
                    messages.error(request, "Avval qaydni bekor qiling.")
                else:
                    x.delete()
                    messages.info(request, "Rasxod o'chirildi.")
        except ValidationError as e:
            messages.error(request, getattr(e, "message", str(e)))
    return redirect("rasxod_list")
