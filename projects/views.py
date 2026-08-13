import datetime
import io
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from django.db.models import Sum

from .models import (_LINE_TOTAL, Firma, GrafikRow, LimitChangeItem,
                     LimitChangeRequest, LimitItem, Project, WeeklyRequest,
                     WeeklyRequestItem, sync_limit_items)
from .roles import (
    can_access_project, is_admin, is_director, is_prorab, is_pto,
    user_firma, visible_firmas, visible_projects,
)


def _to_dec(v):
    """Excel katagidan Decimal (bo'sh/xato bo'lsa None)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _money(value):
    return f"{value:,.0f}".replace(",", " ")


def _firma_yoki_403(request, project):
    """Obyekt foydalanuvchining firmasiga tegishli bo'lmasa 403 (firma izolyatsiyasi)."""
    if not can_access_project(request.user, project):
        raise PermissionDenied("Bu obyekt sizning firmangizga tegishli emas.")
    return project

# eslatma: _qty pastda (bitta joyda) aniqlangan — oldin shu yerda dublikat bor edi


KIND_NOMI = {"material": "Material", "labor": "Ish haqi", "machinery": "Mashina chasti",
             "other": "Ko'zda tutilmagan xarajatlar"}
KINDS = ("material", "labor", "machinery", "other")

# Tasdiqlash zanjirida «jarayonda» statuslar:
# snab (snabjeniye ko'rigida) -> pto2 (PTO xulosasida) -> dir -> adm
LIM_JARAYON = ["snab", "pto2", "dir", "adm"]


def _snab_bor(p):
    """Obyektga snabjeniye xodimi biriktirilganmi?"""
    from .models import UserProfile
    return UserProfile.objects.filter(role="snab", projects=p).exists()


def _limit_boshlangich(p):
    """Yangi limit so'rovining boshlang'ich bosqichi:
    snabjeniye bo'lsa — unga; bo'lmasa to'g'ridan-to'g'ri direktorga."""
    return "snab" if _snab_bor(p) else "dir"


def _limit_yubor_xabar(status):
    return ("Limit snabjeniye ko'rigiga yuborildi. Zanjir: snabjeniye → PTO xulosasi → direktor → admin."
            if status == "snab" else
            "Limit direktor tasdig'iga yuborildi. Direktor, so'ng admin tasdiqlagach kuchga kiradi.")


def _items_by_kind(items):
    """Qatorlarni kategoriya bo'yicha yig'ish: {kind: summa}."""
    d = {k: Decimal("0.00") for k in KINDS}
    for it in items:
        k = it.kind if it.kind in d else "material"
        d[k] += (it.quantity or Decimal("0")) * (it.unit_price or Decimal("0"))
    return d


def _limit_oshish(project, qoshimcha, exclude_request_id=None):
    """Berilgan qatorlar limitdan oshadimi?

    `qoshimcha` — {kind: summa}. Tasdiqlangan sarf hisobga olinadi
    (`exclude_request_id` berilsa, o'sha so'rov sarfdan chiqariladi — qayta
    tasdiqlashda o'zini ikki marta sanamaslik uchun).
    Ro'yxat qaytaradi: har bir oshgan kategoriya bo'yicha tushuntirish.
    Limiti 0 bo'lgan (belgilanmagan) kategoriya tekshirilmaydi.
    """
    qs = WeeklyRequestItem.objects.filter(
        request__project=project,
        request__status=WeeklyRequest.Status.APPROVED,
    )
    if exclude_request_id:
        qs = qs.exclude(request_id=exclude_request_id)
    sarf = {k: Decimal("0.00") for k in KINDS}
    for r in qs.values("kind").annotate(s=Sum(_LINE_TOTAL)):
        if r["kind"] in sarf:
            sarf[r["kind"]] = r["s"] or Decimal("0.00")

    lim = {
        "material": project.limit_material or Decimal("0.00"),
        "labor": project.limit_labor or Decimal("0.00"),
        "machinery": project.limit_machinery or Decimal("0.00"),
        "other": project.limit_other or Decimal("0.00"),
    }
    oshgan = []
    for k in KINDS:
        so_ralgan = qoshimcha.get(k, Decimal("0.00"))
        if so_ralgan <= 0 or lim[k] <= 0:
            continue
        qoldiq = lim[k] - sarf[k]
        if so_ralgan > qoldiq:
            oshgan.append(
                f"{KIND_NOMI[k]}: qoldiq {_money(max(qoldiq, Decimal('0')))}, "
                f"so'ralmoqda {_money(so_ralgan)} "
                f"(oshadi: {_money(so_ralgan - qoldiq)})"
            )
    return oshgan


def _hafta_sana_xatosi(project, ws, we, exclude_id=None):
    """Haftalik so'rov sanalarini tekshiradi. Xato bo'lsa matn, bo'lmasa None qaytaradi."""
    if we < ws:
        return (f"Hafta oxiri ({we:%d.%m.%Y}) hafta boshidan ({ws:%d.%m.%Y}) oldin bo'lishi mumkin emas.")
    if (we - ws).days > 31:
        return (f"Davr juda uzun ({(we - ws).days + 1} kun). Haftalik so'rov 31 kundan oshmasin.")
    qs = project.weekly_requests.filter(week_start=ws, week_end=we)
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    if qs.exists():
        return (f"Bu obyektda {ws:%d.%m.%Y} — {we:%d.%m.%Y} davri uchun so'rov allaqachon mavjud. "
                f"Mavjudini tahrirlang yoki boshqa davr tanlang.")
    return None


def _limitga_yangi_materiallar(req):
    """Haftalik so'rovdagi, umumiy limitda YO'Q materiallarni umumiy limitga qo'shadi.

    Miqdor = 0 → limit SUMMASI oshmaydi, material faqat loyiha ro'yxatiga kiradi.
    Shu sabab `recompute_limits()` CHAQIRILMAYDI: tarkibi bo'sh, lekin limiti
    raqamli kiritilgan obyektlarda u limitni nolga tushirib yuborardi.
    Qo'shilgan nomlar ro'yxatini qaytaradi.
    """
    p = req.project
    mavjud = {(li.kind, (li.name or "").strip().lower()) for li in p.limit_items.all()}
    yangi, korilgan = [], set()
    for it in req.items.all():
        nom = (it.name or "").strip()
        kalit = (it.kind, nom.lower())
        if not nom or kalit in mavjud or kalit in korilgan:
            continue
        korilgan.add(kalit)
        yangi.append(LimitItem(
            project=p, kind=it.kind, name=nom, unit=it.unit or "",
            quantity=Decimal("0"), unit_price=it.unit_price or Decimal("0"),
        ))
    if yangi:
        LimitItem.objects.bulk_create(yangi)
    return [li.name for li in yangi]


RANGLAR = {
    "limitsiz": "#8a8d90",
    "normal": "#107e3e",
    "limitga yaqin": "#e9730c",
    "RUXSAT KERAK": "#bb0000",
}

HOLAT_CLASS = {
    "normal": "ok",
    "limitga yaqin": "warn",
    "RUXSAT KERAK": "bad",
    "limitsiz": "",
}


def _holat(limit, sarf):
    """Limit holatini bir joyda aniqlash (model bilan bir xil mantiq)."""
    if limit <= 0:
        return "limitsiz"
    if sarf > limit:
        return "RUXSAT KERAK"
    if sarf / limit >= Decimal("0.9"):
        return "limitga yaqin"
    return "normal"


@login_required
def dashboard(request):
    firma_id = request.GET.get("firma") or ""
    obyektlar = visible_projects(request.user).order_by("code")
    if firma_id:
        obyektlar = obyektlar.filter(firma_id=firma_id)
    qatorlar = []
    jami_limit = Decimal("0.00")
    jami_mat = Decimal("0.00")
    jami_lab = Decimal("0.00")
    jami_mach = Decimal("0.00")
    jami_oth = Decimal("0.00")
    for p in obyektlar:
        split = p.sarf_by_kind()
        mat, lab, mach, oth = split["material"], split["labor"], split["machinery"], split["other"]
        sarf = mat + lab + mach + oth
        limit = p.budget_total or Decimal("0.00")
        holat = _holat(limit, sarf)
        foiz = float(sarf) / float(limit) * 100 if limit else 0
        jami_limit += limit
        jami_mat += mat
        jami_lab += lab
        jami_mach += mach
        jami_oth += oth
        qatorlar.append({
            "obj": p,
            "limit_str": _money(limit),
            "sarf_str": _money(sarf),
            "mat_str": _money(mat),
            "lab_str": _money(lab),
            "mach_str": _money(mach),
            "oth_str": _money(oth),
            "qoldiq_str": _money(limit - sarf),
            "holat": holat,
            "rang": RANGLAR.get(holat, "#6c757d"),
            "holat_class": HOLAT_CLASS.get(holat, ""),
            "foiz": round(foiz),
            "foiz_bar": min(round(foiz), 100),
        })
    jami_sarf = jami_mat + jami_lab + jami_mach + jami_oth

    from django.db.models import DecimalField, ExpressionWrapper, F, Sum
    _LT = ExpressionWrapper(F("quantity") * F("unit_price"),
                            output_field=DecimalField(max_digits=20, decimal_places=2))

    _lq = visible_projects(request.user)
    if firma_id:
        _lq = _lq.filter(firma_id=firma_id)
    la = _lq.aggregate(m=Sum("limit_material"), l=Sum("limit_labor"),
                       k=Sum("limit_machinery"), o=Sum("limit_other"))
    lim_mat = la["m"] or Decimal("0")
    lim_lab = la["l"] or Decimal("0")
    lim_mach = la["k"] or Decimal("0")
    lim_oth = la["o"] or Decimal("0")

    def _pct(part, whole):
        return round(float(part) / float(whole) * 100) if whole else 0

    foiz_used = _pct(jami_sarf, jami_limit)
    foiz_bar = min(foiz_used, 100)

    tot = jami_limit or Decimal("1")
    c1 = float(lim_mat) / float(tot) * 100
    c2 = c1 + float(lim_lab) / float(tot) * 100
    c3 = c2 + float(lim_mach) / float(tot) * 100
    donut = ("conic-gradient(#0a6ed1 0 %.2f%%, #5899da %.2f%% %.2f%%, "
             "#89c3f0 %.2f%% %.2f%%, #c7d9ee %.2f%% 100%%)"
             % (c1, c1, c2, c2, c3, c3))

    kategoriyalar = [
        {"nom": "Material", "rang": "#2563eb", "limit_str": _money(lim_mat), "sarf_str": _money(jami_mat),
         "pct": _pct(jami_mat, lim_mat), "bar": min(_pct(jami_mat, lim_mat), 100)},
        {"nom": "Ish haqi", "rang": "#60a5fa", "limit_str": _money(lim_lab), "sarf_str": _money(jami_lab),
         "pct": _pct(jami_lab, lim_lab), "bar": min(_pct(jami_lab, lim_lab), 100)},
        {"nom": "Mashina chasti", "rang": "#bfdbfe", "limit_str": _money(lim_mach), "sarf_str": _money(jami_mach),
         "pct": _pct(jami_mach, lim_mach), "bar": min(_pct(jami_mach, lim_mach), 100)},
        {"nom": "Ko'zda tutilmagan", "rang": "#c7d9ee", "limit_str": _money(lim_oth), "sarf_str": _money(jami_oth),
         "pct": _pct(jami_oth, lim_oth), "bar": min(_pct(jami_oth, lim_oth), 100)},
    ]

    # haftalik sarf grafigi (tasdiqlangan)
    _cq = WeeklyRequestItem.objects.filter(
        request__status="approved", request__project__in=visible_projects(request.user))
    if firma_id:
        _cq = _cq.filter(request__project__firma_id=firma_id)
    crows = list(
        _cq.values("request__week_start").annotate(s=Sum(_LT)).order_by("request__week_start")
    )
    W, H = 640, 170
    pts, labels = [], []
    if crows:
        vals = [float(r["s"] or 0) for r in crows]
        labels = [r["request__week_start"].strftime("%d.%m") for r in crows]
        mx = max(vals) or 1
        n = len(vals)
        for i, v in enumerate(vals):
            x = (i / max(n - 1, 1)) * W
            y = H - (v / mx) * (H - 26) - 10
            pts.append((round(x, 1), round(y, 1)))
        if n == 1:
            pts = [(0, pts[0][1]), (W, pts[0][1])]
            labels = [labels[0], labels[0]]
    line_pts = " ".join(f"{x},{y}" for x, y in pts)
    area_pts = (f"0,{H} " + line_pts + f" {W},{H}") if pts else ""

    # amaliy ko'rsatkichlar (jiddiy nazorat)
    from .models import Firma, LimitChangeRequest
    oshgan = sum(1 for q in qatorlar if q["holat"] == "RUXSAT KERAK")
    yaqin = sum(1 for q in qatorlar if q["holat"] == "limitga yaqin")
    limitli = sum(1 for q in qatorlar if q["obj"].budget_total > 0)
    _pq = LimitChangeRequest.objects.filter(
        status__in=LIM_JARAYON, project__in=visible_projects(request.user))
    if firma_id:
        _pq = _pq.filter(project__firma_id=firma_id)
    pending_count = _pq.count()
    qolgan_foiz = max(0, 100 - foiz_used)

    # Chart.js uchun raqamli ma'lumot
    cat_values = [float(lim_mat), float(lim_lab), float(lim_mach), float(lim_oth)]
    wk_labels = [r["request__week_start"].strftime("%d.%m") for r in crows]
    wk_values = [float(r["s"] or 0) for r in crows]

    # --- «Limit kiritish» tab (bitta oynada) ---
    admin_u = request.user.is_superuser
    _fqs = visible_firmas(request.user).order_by("name")
    if firma_id:
        _fqs = _fqs.filter(id=firma_id)
    entry_firms = []
    for f in _fqs:
        _rows = [{"p": pp, "editable": admin_u or pp.budget_total <= 0}
                 for pp in visible_projects(request.user).filter(firma=f).order_by("code")]
        if _rows:
            entry_firms.append({"firma": f, "rows": _rows})
    entry_no_firm = []
    if not firma_id and admin_u:
        entry_no_firm = [{"p": pp, "editable": True}
                         for pp in Project.objects.filter(firma__isnull=True).order_by("code")]
    _all = list(visible_projects(request.user))
    _tab = request.GET.get("tab")
    if _tab == "kiritish" and is_pto(request.user):
        active_tab = "kiritish"
    elif _tab == "tasdiqlar" and request.user.is_superuser:
        active_tab = "tasdiqlar"
    else:
        active_tab = "korish"
    # Tasdiqlar tabi: DIREKTOR o'z navbatini (dir) ko'radi, ADMIN o'z navbatini (adm/submitted).
    # Superuser ikkalasini ham ko'radi (avval direktor navbati).
    from .roles import is_director
    from .roles import is_snab as _is_snab_f
    # Direktor navbati: haqiqiy direktor YOKI ASOSIY admin (admin1/2 emas)
    from .roles import is_asosiy_admin as _asosiy_f
    _dir = (is_director(request.user) and not request.user.is_superuser)         or _asosiy_f(request.user)
    _adm = is_admin(request.user)
    _snb = _is_snab_f(request.user)
    _pto_f = is_pto(request.user)
    tas_lim, tas_wk, tas_lim2, tas_wk2 = [], [], [], []
    tas_lim_s, tas_lim_p2 = [], []
    if _dir:
        tas_lim, tas_wk = _tasdiqlar_data("dir", user=request.user)
    if _adm:
        tas_lim2, tas_wk2 = _tasdiqlar_data("adm", user=request.user)
    if _snb or _adm:
        tas_lim_s, _x = _tasdiqlar_data("snab", user=request.user)
    if _pto_f:
        tas_lim_p2, _x = _tasdiqlar_data("pto2", user=request.user)
    tas_show = _dir or _adm or _snb or bool(tas_lim_p2)
    from django.utils import timezone as _tz
    _h = _tz.localtime().hour
    greeting = "Xayrli tong" if _h < 12 else ("Xayrli kun" if _h < 18 else "Xayrli kech")

    kontekst = {
        "entry_firms": entry_firms,
        "entry_no_firm": entry_no_firm,
        "entry_admin": admin_u,
        "is_pto": is_pto(request.user),
        "jami_obj": len(_all),
        "limitli_obj": sum(1 for pp in _all if pp.budget_total > 0),
        "active_tab": active_tab,
        "tas_lim": tas_lim, "tas_wk": tas_wk,          # direktor navbati
        "tas_lim2": tas_lim2, "tas_wk2": tas_wk2,      # admin navbati
        "tas_lim_s": tas_lim_s,                        # snabjeniye navbati
        "tas_lim_p2": tas_lim_p2,                      # PTO xulosasi navbati
        "is_director": _dir, "is_snab": _snb, "tas_show": tas_show,
        "tas_count": (len(tas_lim) + len(tas_wk) + len(tas_lim2) + len(tas_wk2)
                      + len(tas_lim_s) + len(tas_lim_p2)),
        "greeting": greeting,
        "qatorlar": qatorlar,
        "cat_values": cat_values,
        "wk_labels": wk_labels,
        "wk_values": wk_values,
        "jami_limit_str": _money(jami_limit),
        "jami_sarf_str": _money(jami_sarf),
        "jami_mat_str": _money(jami_mat),
        "jami_lab_str": _money(jami_lab),
        "jami_mach_str": _money(jami_mach),
        "jami_oth_str": _money(jami_oth),
        "jami_qoldiq_str": _money(jami_limit - jami_sarf),
        "soni": len(qatorlar),
        "foiz_used": foiz_used,
        "foiz_bar": foiz_bar,
        "qolgan_foiz": qolgan_foiz,
        "donut": donut,
        "kategoriyalar": kategoriyalar,
        "line_pts": line_pts,
        "area_pts": area_pts,
        "chart_w": W,
        "chart_h": H,
        "chart_labels": labels,
        "obyektlar_soni": len(qatorlar),
        "firmalar_soni": 1 if firma_id else visible_firmas(request.user).count(),
        "limitli": limitli,
        "oshgan": oshgan,
        "yaqin": yaqin,
        "pending_count": pending_count,
        "firmalar": visible_firmas(request.user).order_by("name"),
        "sel_firma": firma_id,
    }
    return render(request, "projects/dashboard.html", kontekst)


@login_required
def firmalar(request):
    """Obyektlar — tablar: «Hammasi» + har bir firma alohida."""
    firmalar_data = []
    all_objs = []
    tot_limit = Decimal("0")
    tot_sarf = Decimal("0")
    for f in visible_firmas(request.user).order_by("name"):
        objs = []
        f_limit = Decimal("0")
        f_sarf = Decimal("0")
        for pr in visible_projects(request.user).filter(firma=f).order_by("code"):
            bt = pr.budget_total
            srf = pr.sarflangan()
            f_limit += bt
            f_sarf += srf
            foiz = round(float(srf) / float(bt) * 100) if bt else 0
            row = {
                "obj": pr,
                "firma_name": f.name,
                "status": pr.get_status_display(),
                "status_key": pr.status,
                "location": pr.location or "",
                "start": pr.start_date,
                "end": pr.end_date,
                "has_limit": bt > 0,
                "limit_str": _money(bt),
                "sarf_str": _money(srf),
                "qoldiq_str": _money(bt - srf),
                "foiz": foiz,
                "bar": min(foiz, 100),
                "rang": "#ba1a1a" if foiz > 100 else ("#b45309" if foiz >= 90 else "#2563eb"),
            }
            objs.append(row)
            all_objs.append(row)
        tot_limit += f_limit
        tot_sarf += f_sarf
        firmalar_data.append({
            "obj": f,
            "objects": objs,
            "count": len(objs),
            "limit_str": _money(f_limit),
            "sarf_str": _money(f_sarf),
            "qoldiq_str": _money(f_limit - f_sarf),
            "foiz": round(float(f_sarf) / float(f_limit) * 100) if f_limit else 0,
        })
    hammasi = {
        "objects": all_objs,
        "count": len(all_objs),
        "limit_str": _money(tot_limit),
        "sarf_str": _money(tot_sarf),
        "qoldiq_str": _money(tot_limit - tot_sarf),
    }
    return render(request, "projects/firmalar.html", {
        "firmalar": firmalar_data,
        "hammasi": hammasi,
        "soni": len(firmalar_data),
    })


def _qty(x):
    """Miqdorni chiroyli format qilish (3 xonagacha, ortiqcha nol tashlanadi)."""
    x = Decimal(x or 0)
    s = f"{x:,.3f}".replace(",", " ")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"


@login_required
def m29(request):
    """M-29 — material sarfining reja-fakt tahlili.

    Reja  = tasdiqlangan haftalik so'rovlardagi material qatorlari.
    Fakt  = ombordan chiqim (rasxod) — StockMovement OUT.
    Material nomi + o'lchov birligi bo'yicha birlashtiriladi.
    """
    from django.db.models import DecimalField, ExpressionWrapper, F, Sum
    from ombor.models import StockMovement

    _LT = ExpressionWrapper(F("quantity") * F("unit_price"),
                            output_field=DecimalField(max_digits=20, decimal_places=2))

    firma_id = request.GET.get("firma") or ""
    obj_id = request.GET.get("obyekt") or ""
    d1 = request.GET.get("d1") or ""
    d2 = request.GET.get("d2") or ""

    projs = visible_projects(request.user).order_by("code")
    if firma_id:
        projs = projs.filter(firma_id=firma_id)

    # --- REJA: tasdiqlangan haftalik material qatorlari ---
    reja_q = WeeklyRequestItem.objects.filter(
        request__status=WeeklyRequest.Status.APPROVED,
        kind=WeeklyRequestItem.Kind.MATERIAL,
        request__project__in=visible_projects(request.user),
    )
    if firma_id:
        reja_q = reja_q.filter(request__project__firma_id=firma_id)
    if obj_id:
        reja_q = reja_q.filter(request__project_id=obj_id)
    if d1:
        reja_q = reja_q.filter(request__week_start__gte=d1)
    if d2:
        reja_q = reja_q.filter(request__week_start__lte=d2)
    reja_rows = reja_q.values("name", "unit").annotate(qty=Sum("quantity"), summa=Sum(_LT))

    # --- FAKT: ombor chiqimi (OUT — miqdor/summa manfiy ishorali) ---
    fakt_q = StockMovement.objects.filter(
        direction=StockMovement.OUT, project__in=visible_projects(request.user))
    if firma_id:
        fakt_q = fakt_q.filter(project__firma_id=firma_id)
    if obj_id:
        fakt_q = fakt_q.filter(project_id=obj_id)
    if d1:
        fakt_q = fakt_q.filter(date__gte=d1)
    if d2:
        fakt_q = fakt_q.filter(date__lte=d2)
    fakt_rows = fakt_q.values("material__name", "material__unit").annotate(
        qty=Sum("quantity"), summa=Sum("total_cost"))

    def _key(n, u):
        return ((n or "").strip().lower(), (u or "").strip().lower())

    data = {}

    def _slot(name, unit):
        k = _key(name, unit)
        if k not in data:
            data[k] = {"name": name or "—", "unit": unit or "—",
                       "reja_qty": Decimal("0"), "reja_sum": Decimal("0"),
                       "fakt_qty": Decimal("0"), "fakt_sum": Decimal("0")}
        return data[k]

    for r in reja_rows:
        s = _slot(r["name"], r["unit"])
        s["reja_qty"] += r["qty"] or Decimal("0")
        s["reja_sum"] += r["summa"] or Decimal("0")
    for r in fakt_rows:
        s = _slot(r["material__name"], r["material__unit"])
        s["fakt_qty"] += abs(r["qty"] or Decimal("0"))
        s["fakt_sum"] += abs(r["summa"] or Decimal("0"))

    rows = []
    t_reja = Decimal("0")
    t_fakt = Decimal("0")
    n_perer = 0
    n_ekon = 0
    for k in sorted(data.keys()):
        d = data[k]
        farq = d["fakt_sum"] - d["reja_sum"]
        if d["reja_sum"] == 0 and d["fakt_sum"] > 0:
            holat, rang = "Rejasiz sarf", "#b91c1c"
            n_perer += 1
        elif farq > 0:
            holat, rang = "Pererasxod", "#b91c1c"
            n_perer += 1
        elif farq < 0:
            holat, rang = "Ekonomiya", "#15803d"
            n_ekon += 1
        else:
            holat, rang = "Norma", "#64748b"
        t_reja += d["reja_sum"]
        t_fakt += d["fakt_sum"]
        rows.append({
            "name": d["name"],
            "unit": d["unit"],
            "reja_qty": _qty(d["reja_qty"]),
            "reja_sum": _money(d["reja_sum"]),
            "fakt_qty": _qty(d["fakt_qty"]),
            "fakt_sum": _money(d["fakt_sum"]),
            "farq_sum": _money(abs(farq)),
            "farq_neg": farq < 0,
            "farq_zero": farq == 0,
            "holat": holat,
            "rang": rang,
        })

    farq_jami = t_fakt - t_reja
    sel_obj = projs.filter(id=obj_id).first() if obj_id else None

    kontekst = {
        "rows": rows,
        "soni": len(rows),
        "reja_jami": _money(t_reja),
        "fakt_jami": _money(t_fakt),
        "farq_jami": _money(abs(farq_jami)),
        "farq_neg": farq_jami < 0,
        "farq_zero": farq_jami == 0,
        "n_perer": n_perer,
        "n_ekon": n_ekon,
        "firmalar": visible_firmas(request.user).order_by("name"),
        "obyektlar": projs,
        "sel_firma": firma_id,
        "sel_obj": obj_id,
        "sel_obj_name": (f"{sel_obj.code} — {sel_obj.name}") if sel_obj else "",
        "d1": d1,
        "d2": d2,
    }
    return render(request, "projects/m29.html", kontekst)


@login_required
def virtual_ofis(request):
    """Virtual ofis — har biri o'z vazifasiga ega avtomat agentlar.
    Har agent ERP ma'lumotini skanerlaydi va vazifalar/ogohlantirishlar chiqaradi."""
    from django.utils import timezone
    from django.urls import reverse as _rev
    from ombor.models import StockBalance
    today = timezone.localdate()

    agents = []

    def _u_obj(pid):
        return _rev("project_detail", args=[pid])

    # 1) Limit nazoratchisi
    tasks = []
    for p in visible_projects(request.user).select_related("firma"):
        lim = p.budget_total
        if lim <= 0:
            continue
        sarf = p.sarflangan()
        if sarf > lim:
            tasks.append({"text": f"{p.code} — {p.name}: limitdan oshgan ({_money(sarf)} / {_money(lim)})",
                          "level": "bad", "url": _u_obj(p.id)})
        elif float(sarf) / float(lim) >= 0.9:
            tasks.append({"text": f"{p.code} — {p.name}: limitga yaqin ({round(float(sarf)/float(lim)*100)}%)",
                          "level": "warn", "url": _u_obj(p.id)})
    agents.append({"name": "Limit nazoratchisi", "role": "Obyekt limitlarini kuzatadi",
                   "av": "L", "tasks": tasks})

    # 2) Tasdiq kotibi
    tasks = []
    _vp = visible_projects(request.user)
    lc = LimitChangeRequest.objects.filter(status__in=LIM_JARAYON, project__in=_vp).count()
    wc = WeeklyRequest.objects.filter(status="submitted", project__in=_vp).count()
    if lc:
        tasks.append({"text": f"{lc} ta limit o'zgartirish so'rovi tasdiq kutmoqda",
                      "level": "warn", "url": _rev("tasdiqlar")})
    if wc:
        tasks.append({"text": f"{wc} ta haftalik so'rov tasdiq kutmoqda",
                      "level": "warn", "url": _rev("tasdiqlar")})
    agents.append({"name": "Tasdiq kotibi", "role": "Tasdiq kutayotgan so'rovlarni yig'adi",
                   "av": "T", "tasks": tasks})

    # 3) Ombor nazoratchisi
    tasks = []
    low = (StockBalance.objects.select_related("material", "warehouse")
           .filter(quantity__lte=0))
    if not request.user.is_superuser:
        low = low.filter(warehouse__project__in=visible_projects(request.user))
    low = low.order_by("warehouse__name")[:30]
    for b in low:
        tasks.append({"text": f"{b.material.name} — {b.warehouse.kod_nom}: qoldiq {_qty(b.quantity)} (tugagan)",
                      "level": "bad", "url": _rev("ombor")})
    agents.append({"name": "Ombor nazoratchisi", "role": "Tugagan/kam material qoldiqlarini topadi",
                   "av": "O", "tasks": tasks})

    # 4) Muddat agenti
    tasks = []
    for p in visible_projects(request.user).filter(status="active").exclude(end_date__isnull=True):
        if p.end_date < today:
            tasks.append({"text": f"{p.code} — {p.name}: muddati o'tgan ({p.end_date:%d.%m.%Y})",
                          "level": "bad", "url": _u_obj(p.id)})
        elif (p.end_date - today).days <= 14:
            tasks.append({"text": f"{p.code} — {p.name}: {(p.end_date - today).days} kun qoldi ({p.end_date:%d.%m.%Y})",
                          "level": "warn", "url": _u_obj(p.id)})
    agents.append({"name": "Muddat agenti", "role": "Obyekt muddatlarini nazorat qiladi",
                   "av": "M", "tasks": tasks})

    # 5) Sarf tahlilchisi (kategoriya bo'yicha pererasxod)
    tasks = []
    KAT = [("Material", "limit_material", "material"), ("Ish haqi", "limit_labor", "labor"),
           ("Mashina chasti", "limit_machinery", "machinery"),
           ("Ko'zda tutilmagan", "limit_other", "other")]
    for p in visible_projects(request.user):
        if p.budget_total <= 0:
            continue
        split = p.sarf_by_kind()
        for nom, lf, key in KAT:
            klim = getattr(p, lf) or Decimal("0")
            ksarf = split[key]
            if klim > 0 and ksarf > klim:
                tasks.append({"text": f"{p.code}: «{nom}» pererasxod ({_money(ksarf)} / {_money(klim)})",
                              "level": "bad", "url": _u_obj(p.id)})
    agents.append({"name": "Sarf tahlilchisi", "role": "Kategoriya bo'yicha pererasxodni aniqlaydi",
                   "av": "S", "tasks": tasks})

    # umumiy holat
    for a in agents:
        a["count"] = len(a["tasks"])
        a["bad"] = sum(1 for t in a["tasks"] if t["level"] == "bad")
        a["ok"] = a["count"] == 0
    jami_vazifa = sum(a["count"] for a in agents)
    band_agent = sum(1 for a in agents if a["count"] > 0)

    return render(request, "projects/virtual_ofis.html", {
        "agents": agents, "jami_vazifa": jami_vazifa,
        "band_agent": band_agent, "agent_soni": len(agents),
    })


@login_required
def haftalik_tarix(request):
    """Tasdiqdan o'tgan (eski) haftalik limitlar tarixi — filtr bilan ko'rish."""
    firma_id = request.GET.get("firma") or ""
    obj_id = request.GET.get("obyekt") or ""
    d1 = request.GET.get("d1") or ""
    d2 = request.GET.get("d2") or ""
    holat = request.GET.get("holat") or "approved"  # approved / submitted / draft / all

    projs = visible_projects(request.user).order_by("code")
    if firma_id:
        projs = projs.filter(firma_id=firma_id)

    qs = WeeklyRequest.objects.filter(
        project__in=visible_projects(request.user)).select_related(
        "project", "project__firma", "created_by", "approved_by").prefetch_related("items")
    if holat in ("approved", "submitted", "draft"):
        qs = qs.filter(status=holat)
    if firma_id:
        qs = qs.filter(project__firma_id=firma_id)
    if obj_id:
        qs = qs.filter(project_id=obj_id)
    if d1:
        qs = qs.filter(week_start__gte=d1)
    if d2:
        qs = qs.filter(week_start__lte=d2)
    qs = qs.order_by("-week_start", "-id")

    rows = []
    t_jami = Decimal("0")
    for w in qs:
        s = {k: Decimal("0") for k in KINDS}
        cnt = 0
        for it in w.items.all():
            s[it.kind] = s.get(it.kind, Decimal("0")) + it.total
            cnt += 1
        jami = s["material"] + s["labor"] + s["machinery"] + s["other"]
        t_jami += jami
        rows.append({
            "id": w.id, "obj": w.project,
            "firma": w.project.firma.name if w.project.firma_id else "—",
            "week_start": w.week_start, "week_end": w.week_end, "number": w.number,
            "note": w.note,
            "kiritdi": w.created_by.username if w.created_by_id else "—",
            "kiritilgan": w.created_at,
            "tasdiqladi": w.approved_by.username if w.approved_by_id else "",
            "tasdiq_sana": w.approved_at,
            "is_approved": w.status == "approved",
            "mat_str": _money(s["material"]), "lab_str": _money(s["labor"]),
            "mach_str": _money(s["machinery"]), "oth_str": _money(s["other"]),
            "jami_str": _money(jami), "soni": cnt,
        })

    sel_obj = projs.filter(id=obj_id).first() if obj_id else None
    kontekst = {
        "rows": rows, "soni": len(rows), "jami_str": _money(t_jami),
        "firmalar": visible_firmas(request.user).order_by("name"), "obyektlar": projs,
        "sel_firma": firma_id, "sel_obj": obj_id,
        "sel_obj_name": (f"{sel_obj.code} — {sel_obj.name}") if sel_obj else "",
        "d1": d1, "d2": d2, "holat": holat,
    }
    return render(request, "projects/haftalik_tarix.html", kontekst)


@login_required
def project_detail(request, pk):
    p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))

    pending = p.limit_requests.filter(status__in=LIM_JARAYON).first()
    # Oxirgi rad etilgan so'rov — PTO sababini ko'rsin (kutilayotgani bo'lmasa)
    rad_sorov = None
    if not pending:
        rad_sorov = (p.limit_requests.filter(status=LimitChangeRequest.Status.REJECTED)
                     .select_related("decided_by").order_by("-decided_at").first())

    # «Tasdiqlandi» xabari — DOIMIY ko'rinadi (foydalanuvchi xohishi: OK tugmasisiz).
    # Uyulib ketmasligi uchun faqat ENG OXIRGI tasdiq ko'rsatiladi.
    tasdiq_wk = (p.weekly_requests
                 .filter(status="approved", created_by=request.user)
                 .select_related("approved_by").order_by("-approved_at")[:1])
    tasdiq_lim = (p.limit_requests
                  .filter(status=LimitChangeRequest.Status.APPROVED,
                          requested_by=request.user)
                  .select_related("decided_by").order_by("-decided_at").first())

    # Limit blokidagi DOIMIY yashil holat — oxirgi tasdiqlangan so'rov (hammaga ko'rinadi)
    oxirgi_tasdiq = None
    if not pending:
        oxirgi_tasdiq = (p.limit_requests
                         .filter(status=LimitChangeRequest.Status.APPROVED)
                         .select_related("decided_by").order_by("-decided_at").first())

    split = p.sarf_by_kind()
    sarf = split["material"] + split["labor"] + split["machinery"] + split["other"]
    limit = p.budget_total or Decimal("0.00")
    holat = _holat(limit, sarf)
    foiz = float(sarf) / float(limit) * 100 if limit else 0

    # kategoriyalar: limit vs sarf (material / ish haqi / mashina / ko'zda tutilmagan)
    kats = [
        ("Material", p.limit_material or Decimal("0"), split["material"], "mat"),
        ("Ish haqi", p.limit_labor or Decimal("0"), split["labor"], "lab"),
        ("Mashina chasti", p.limit_machinery or Decimal("0"), split["machinery"], "mach"),
        ("Ko'zda tutilmagan", p.limit_other or Decimal("0"), split["other"], "oth"),
    ]
    kategoriyalar = []
    for nom, klimit, ksarf, cls in kats:
        kfoiz = round(float(ksarf) / float(klimit) * 100) if klimit else 0
        kategoriyalar.append({
            "nom": nom, "cls": cls,
            "limit_str": _money(klimit), "sarf_str": _money(ksarf),
            "qoldiq_str": _money(klimit - ksarf),
            "foiz": kfoiz, "foiz_bar": min(kfoiz, 100),
            "oshgan": ksarf > klimit and klimit > 0,
        })

    # Nom -> birlik xaritasi (material birligini avtomatik to'ldirish uchun)
    from ombor.models import Material
    unit_map = {}

    def _umap(nm, un):
        k = (nm or "").strip().lower()
        if k and un and k not in unit_map:
            unit_map[k] = un
    for m in Material.objects.all():
        _umap(m.name, m.unit)
    for it in WeeklyRequestItem.objects.filter(kind="material").exclude(unit="").values("name", "unit"):
        _umap(it["name"], it["unit"])
    for it in LimitItem.objects.filter(kind="material").exclude(unit="").values("name", "unit"):
        _umap(it["name"], it["unit"])
    mat_names = sorted({m.name for m in Material.objects.all()})

    # Umumiy limit ichi (tarkibi) — BO'LIM bo'yicha ketma-ket guruhlanadi
    # (bo'lim birinchi uchragan tartibda; bo'limsiz qatorlar oxirida)
    LI_CLS = {"material": "mat", "labor": "lab", "machinery": "mach", "other": "oth"}
    li_cat = {k: Decimal("0") for k in KINDS}
    _guruhlar = {}   # bolim -> {"bolim", "masul", "items", "_sum"}
    for it in p.limit_items.all().order_by("id"):
        li_cat[it.kind] += it.total
        qator = {
            "kind": it.kind, "kind_disp": it.get_kind_display(), "cls": LI_CLS.get(it.kind, "mat"),
            "name": it.name, "unit": it.unit,
            "quantity": it.quantity, "price_raw": it.unit_price,
            "qty_str": _qty(it.quantity), "price_str": _money(it.unit_price),
            "sum_str": _money(it.total), "note": it.note,
            "bolim": it.bolim, "masul": it.masul,
            "qoshilgan": it.created_at, "ozgartirilgan": it.updated_at,
            # qo'shilgandan keyin o'zgartirilganmi (1 daqiqadan ortiq farq bo'lsa)
            "ozgargan": bool(it.created_at and it.updated_at
                             and (it.updated_at - it.created_at).total_seconds() > 60),
        }
        kalit = (it.bolim or "").strip()
        g = _guruhlar.get(kalit)
        if g is None:
            g = _guruhlar[kalit] = {"bolim": kalit, "masul": (it.masul or "").strip(),
                                    "items": [], "_sum": Decimal("0")}
        if not g["masul"] and (it.masul or "").strip():
            g["masul"] = it.masul.strip()
        g["items"].append(qator)
        g["_sum"] += it.total
    # bo'limsiz guruh («») oxiriga o'tadi
    limit_groups = [g for k, g in _guruhlar.items() if k] + \
                   [g for k, g in _guruhlar.items() if not k]
    for g in limit_groups:
        g["sum_str"] = _money(g["_sum"])
    # tekis ro'yxat (tahrirlash jadvali uchun) — bo'lim qatorlari yonma-yon turadi
    limit_items = [q for g in limit_groups for q in g["items"]]
    # Har pozitsiyaga KETMA-KET tartib raqami (guruhlar bo'ylab uzluksiz)
    for _nr, q in enumerate(limit_items, start=1):
        q["nr"] = _nr
    bolim_bor = any(g["bolim"] for g in limit_groups)

    # Toifa bo'yicha alohida summa itogolari (jadval ostida ko'rsatiladi)
    li_cat_sums = [
        {"kind": k, "cls": LI_CLS[k], "nom": KIND_NOMI[k], "sum_str": _money(li_cat[k])}
        for k in KINDS
    ]
    li_jami_str = _money(sum(li_cat.values()))

    # Haftalik so'rovda limit tarkibidan tanlash uchun: nom -> {turi, birlik, narx}
    limit_item_map = {}
    for it in limit_items:
        k = (it["name"] or "").strip().lower()
        if k and k not in limit_item_map:
            limit_item_map[k] = {"kind": it["kind"], "unit": it["unit"], "price": float(it["price_raw"])}
    # Loyihaning O'Z materiallari (umumiy limit tarkibi) — haftalik tanlovda birinchi turadi
    proj_names = sorted({it["name"] for it in limit_items if it["name"]})
    mat_names = sorted(set(mat_names) | set(proj_names))

    sorovlar = []
    requests = (
        p.weekly_requests
        .select_related("created_by", "approved_by")
        .prefetch_related("items", "items__work_section")
        .order_by("-week_start", "-id")
    )
    KIND_CLS = {"labor": "lab", "machinery": "mach", "material": "mat", "other": "oth"}
    for r in requests:
        qatorlar = []
        r_sum = {k: Decimal("0") for k in KINDS}
        for it in r.items.all():
            summa = it.total
            r_sum[it.kind] = r_sum.get(it.kind, Decimal("0")) + summa
            qatorlar.append({
                "kind": it.get_kind_display(),
                "cls": KIND_CLS.get(it.kind, "mat"),
                "bolim": it.work_section.name if it.work_section_id else "—",
                "nomi": it.name,
                "birlik": it.unit,
                "miqdor": _qty(it.quantity),
                "narx_str": _money(it.unit_price),
                "summa_str": _money(summa),
                "izoh": it.note,
                "sana": it.created_at,
            })
        sorovlar.append({
            "obj": r,
            "_jami": sum(r_sum.values()),
            "jami_str": _money(sum(r_sum.values())),
            "mat_str": _money(r_sum["material"]),
            "lab_str": _money(r_sum["labor"]),
            "mach_str": _money(r_sum["machinery"]),
            "oth_str": _money(r_sum["other"]),
            "is_approved": r.status == "approved",
            "is_dir": r.status == "dir",            # direktor tasdig'ida
            "is_submitted": r.status == "submitted",  # admin tasdig'ida (direktordan o'tgan)
            "status_disp": r.get_status_display(),
            "qatorlar": qatorlar,
            # Tasdiqlanmagan so'rov limitdan oshadimi (admin oldindan ko'rsin)
            "oshgan": [] if r.status == "approved" else _limit_oshish(
                p, _items_by_kind(r.items.all()), exclude_request_id=r.id),
            # «eski» pastda aniqlanadi — eng oxirgi (yangi) haftadan boshqasi
            "is_old": False,
        })

    # Faqat ENG OXIRGI (eng yangi) haftalik so'rov ochiq turadi; qolganlari «o'tgan» —
    # standart holatda yashiriladi, «Ko'rsatish» bilan ochiladi.
    if sorovlar:
        _yangi = max(sorovlar, key=lambda s: (s["obj"].week_start, s["obj"].id))
        for s in sorovlar:
            s["is_old"] = (s is not _yangi)

    # Yuguruvchi qoldiq + ketma-ket raqam: eng eski haftadan boshlab
    running = Decimal("0")
    qoldiq_map = {}
    seq_map = {}
    for i, s in enumerate(sorted(sorovlar, key=lambda s: (s["obj"].week_start, s["obj"].id)), start=1):
        seq_map[s["obj"].id] = i
        if s["is_approved"]:
            running += s["_jami"]
        qoldiq_map[s["obj"].id] = limit - running
    for s in sorovlar:
        s["seq"] = seq_map[s["obj"].id]
        q = qoldiq_map[s["obj"].id]
        s["qoldiq_str"] = _money(q)
        s["qoldiq_neg"] = q < 0

    # Oxirgi (eng yangi) TASDIQLANGAN haftalik so'rov — limit bo'limida yaqqol ko'rsatiladi.
    # sorovlar «-week_start» tartibda, shuning uchun birinchi approved = eng oxirgisi.
    oxirgi_haftalik = next((s for s in sorovlar if s["is_approved"]), None)

    kontekst = {
        "p": p,
        "limit_str": _money(limit),
        "sarf_str": _money(sarf),
        "qoldiq_str": _money(limit - sarf),
        "oxirgi_haftalik": oxirgi_haftalik,
        "holat": holat,
        "rang": RANGLAR.get(holat, "#6c757d"),
        "foiz": round(foiz),
        "foiz_bar": min(round(foiz), 100),
        "kategoriyalar": kategoriyalar,
        "sorovlar": sorovlar,
        "soni": len(sorovlar),
        "proj_names": proj_names,
        "rad_sorov": rad_sorov,
        "tasdiq_wk": tasdiq_wk,
        "tasdiq_lim": tasdiq_lim,
        "oxirgi_tasdiq": oxirgi_tasdiq,
        "eski_soni": sum(1 for s in sorovlar if s["is_old"]),
        "is_pto": is_pto(request.user),
        "is_admin": is_admin(request.user),
        "is_director": is_director(request.user),
        "limit_set": bool(limit and limit > 0),
        "limit_items": limit_items,
        "limit_groups": limit_groups,
        "bolim_bor": bolim_bor,
        "li_cat_sums": li_cat_sums,
        "li_jami_str": li_jami_str,
        "has_items": bool(limit_items),
        "unit_map": unit_map,
        "mat_names": mat_names,
        "limit_item_map": limit_item_map,
        "pending": pending,
        "pending_new_str": _money(pending.new_total) if pending else "",
        "qoldiq_json": {
            "total": float(limit - sarf),
            "material": float((p.limit_material or Decimal("0")) - split["material"]),
            "labor": float((p.limit_labor or Decimal("0")) - split["labor"]),
            "machinery": float((p.limit_machinery or Decimal("0")) - split["machinery"]),
            "other": float((p.limit_other or Decimal("0")) - split["other"]),
        },
        "rem_mat_str": _money((p.limit_material or Decimal("0")) - split["material"]),
        "rem_lab_str": _money((p.limit_labor or Decimal("0")) - split["labor"]),
        "rem_mach_str": _money((p.limit_machinery or Decimal("0")) - split["machinery"]),
        "rem_oth_str": _money((p.limit_other or Decimal("0")) - split["other"]),
    }
    return render(request, "projects/detail.html", kontekst)


@login_required
def limit_edit(request, pk):
    """PTO limit kiritadi. Limit allaqachon bo'lsa — o'zgartirishga so'rov (admin tasdiqlaydi)."""
    p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))
    if not is_pto(request.user):
        raise PermissionDenied("Limitni faqat PTO kiritadi.")
    if request.method != "POST":
        return redirect("project_detail", pk=pk)

    mat = _to_dec(request.POST.get("material"))
    lab = _to_dec(request.POST.get("labor"))
    mach = _to_dec(request.POST.get("machinery"))
    oth = _to_dec(request.POST.get("other")) or Decimal("0")
    sabab = (request.POST.get("reason") or "").strip()
    if None in (mat, lab, mach) or mat < 0 or lab < 0 or mach < 0 or oth < 0:
        messages.error(request, "Limit noto'g'ri kiritildi.")
        return redirect("project_detail", pk=pk)

    eski = (p.limit_material, p.limit_labor, p.limit_machinery, p.limit_other)
    yangi = (mat, lab, mach, oth)
    if is_admin(request.user):
        # Admin — to'g'ridan-to'g'ri (u tasdiqlovchi)
        p.limit_material, p.limit_labor, p.limit_machinery, p.limit_other = yangi
        p.save(update_fields=["limit_material", "limit_labor", "limit_machinery", "limit_other"])
        messages.success(request, f"Limit belgilandi: {_money(mat + lab + mach + oth)}")
    elif yangi == eski:
        messages.info(request, "Limit o'zgarmadi.")
    else:
        # mavjud limitni o'zgartirish — admin tasdig'i kerak
        pending = p.limit_requests.filter(status__in=LIM_JARAYON).first()
        if pending:
            messages.error(request, "Bu obyekt bo'yicha allaqachon tasdiq kutilayotgan so'rov bor.")
        else:
            _st = _limit_boshlangich(p)
            LimitChangeRequest.objects.create(
                project=p,
                old_material=p.limit_material, old_labor=p.limit_labor,
                old_machinery=p.limit_machinery, old_other=p.limit_other,
                new_material=mat, new_labor=lab, new_machinery=mach, new_other=oth,
                reason=sabab, requested_by=request.user, status=_st,
            )
            messages.success(request, _limit_yubor_xabar(_st))
    return redirect("project_detail", pk=pk)


@login_required
def limit_items_edit(request, pk):
    """PTO umumiy limit ichini (tarkibini) kiritadi/o'zgartiradi.
    Limit yo'q bo'lsa — to'g'ridan-to'g'ri. Bor bo'lsa — admin tasdig'i kerak."""
    p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))
    if not is_pto(request.user):
        raise PermissionDenied("Limit ichini faqat PTO kiritadi.")
    if request.method != "POST":
        return redirect("project_detail", pk=pk)

    kinds = request.POST.getlist("kind")
    names = request.POST.getlist("name")
    units = request.POST.getlist("unit")
    qtys = request.POST.getlist("quantity")
    prices = request.POST.getlist("unit_price")
    inotes = request.POST.getlist("item_note")
    bolims = request.POST.getlist("item_bolim")
    masuls = request.POST.getlist("item_masul")

    valid = set(KINDS)
    items = []
    sums = {k: Decimal("0") for k in KINDS}
    for i in range(len(names)):
        nm = (names[i] or "").strip()
        if not nm:
            continue
        kind = kinds[i] if i < len(kinds) and kinds[i] in valid else "material"
        unit = (units[i] if i < len(units) else "").strip()
        q = _to_dec(qtys[i] if i < len(qtys) else "0") or Decimal("0")
        pr = _to_dec(prices[i] if i < len(prices) else "0") or Decimal("0")
        if q < 0 or pr < 0:
            continue
        items.append({"kind": kind, "name": nm, "unit": unit, "quantity": q, "unit_price": pr,
                      "note": " ".join((inotes[i] if i < len(inotes) else "").split())[:500],
                      "bolim": " ".join((bolims[i] if i < len(bolims) else "").split())[:200],
                      "masul": " ".join((masuls[i] if i < len(masuls) else "").split())[:120]})
        sums[kind] += (q * pr).quantize(Decimal("0.01"))

    if not items:
        messages.error(request, "Kamida bitta tarkib qatorini kiriting.")
        return redirect("project_detail", pk=pk)

    if is_admin(request.user):
        # Asosiy admin — to'g'ridan-to'g'ri qo'llanadi (u tasdiqlovchi)
        with transaction.atomic():
            # o'chirib-qayta yaratmaymiz — o'zgarmagan qatorlarning sanasi saqlansin
            sync_limit_items(p, items)
            p.recompute_limits()
        messages.success(request, f"Limit saqlandi (admin). Umumiy limit: {_money(p.budget_total)}")
    else:
        # PTO limitni kiritadi/o'zgartiradi — HAR DOIM admin tasdig'i kerak (birinchisi ham)
        if p.limit_requests.filter(status__in=LIM_JARAYON).exists():
            messages.error(request, "Tasdiq kutilayotgan so'rov bor — yangi so'rov yuborib bo'lmaydi.")
            return redirect("project_detail", pk=pk)
        reason = (request.POST.get("reason") or "").strip()
        _st = _limit_boshlangich(p)
        with transaction.atomic():
            req = LimitChangeRequest.objects.create(
                project=p,
                old_material=p.limit_material, old_labor=p.limit_labor,
                old_machinery=p.limit_machinery, old_other=p.limit_other,
                new_material=sums["material"], new_labor=sums["labor"],
                new_machinery=sums["machinery"], new_other=sums["other"],
                reason=reason, requested_by=request.user, status=_st,
            )
            LimitChangeItem.objects.bulk_create([LimitChangeItem(request=req, **it) for it in items])
        messages.success(request, _limit_yubor_xabar(_st))
    return redirect("project_detail", pk=pk)


def _qayt_kod_yubor(request, tur, pk, obj_nomi, sabab):
    """PTOga qaytarish uchun 6 xonali tasdiqlash kodini Telegramga yuboradi.
    Kod sessiyada saqlanadi (10 daqiqa, 3 urinish)."""
    import secrets
    import time
    kod = "".join(str(secrets.randbelow(10)) for _ in range(6))
    request.session["qayt_kod"] = {"kod": kod, "tur": tur, "pk": pk,
                                   "sabab": sabab, "ts": time.time(), "urin": 0}
    from .auth2fa import admin_chatlar, tg_send
    prof = getattr(request.user, "profile", None)
    chat = (getattr(prof, "telegram_chat_id", "") or "") if prof else ""
    chats = [chat] if chat else list(admin_chatlar())
    matn = (f"🔐 Tasdiqlash kodi: {kod}\n"
            f"{obj_nomi} — {tur}ni PTOga qaytarish.\n"
            f"So'radi: {request.user.username}. Kod 10 daqiqa amal qiladi.")
    yubordi = False
    for c in chats:
        try:
            if tg_send(c, matn):
                yubordi = True
        except Exception:
            pass
    return yubordi


def _qayt_kod_tekshir(request, tur, pk, kod):
    """Sessiyadagi kod bilan solishtiradi. (sabab, xato) qaytaradi."""
    import time
    d = request.session.get("qayt_kod") or {}
    if d.get("tur") != tur or d.get("pk") != pk:
        return None, "Kod so'ralmagan — jarayonni qaytadan boshlang."
    if time.time() - d.get("ts", 0) > 600:
        request.session.pop("qayt_kod", None)
        return None, "Kod muddati o'tdi (10 daqiqa) — qaytadan boshlang."
    if d.get("urin", 0) >= 3:
        request.session.pop("qayt_kod", None)
        return None, "3 marta xato kiritildi — qaytadan boshlang."
    import secrets as _sec
    if not _sec.compare_digest(kod, d.get("kod", "")):
        d["urin"] = d.get("urin", 0) + 1
        request.session["qayt_kod"] = d
        return None, "Kod noto'g'ri — Telegramdagi kodni tekshirib qayta kiriting."
    sabab = d.get("sabab", "")
    request.session.pop("qayt_kod", None)
    return sabab, ""


@login_required
def limit_return_pto(request, pk):
    """ADMIN tasdiqlangan limitni PTOga QAYTARADI (kamchilik topilganda) —
    Telegram tasdiqlash KODI bilan. Joriy limit tarkibidan «PTO xulosasi»
    bosqichida yangi so'rov ochiladi; amaldagi limit yangi tasdiqqacha o'zgarmaydi."""
    p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))
    if not is_admin(request.user):
        raise PermissionDenied("Limitni faqat admin qaytaradi.")
    if request.method != "POST":
        return redirect("project_detail", pk=pk)
    if p.limit_requests.filter(status__in=LIM_JARAYON).exists():
        messages.error(request, "Bu obyektda tasdiq kutilayotgan so'rov bor — avval uni yakunlang.")
        return redirect("project_detail", pk=pk)
    items = list(p.limit_items.all().order_by("id"))
    if not items:
        messages.error(request, "Limit tarkibi bo'sh — qaytaradigan narsa yo'q.")
        return redirect("project_detail", pk=pk)

    kod = (request.POST.get("kod") or "").strip()
    if not kod:
        # 1-bosqich: sabab keldi -> kod yuborib, tasdiqlash sahifasini ochamiz
        sabab = (request.POST.get("sabab") or "").strip()
        if not sabab:
            messages.error(request, "Qaytarish sababini (kamchilikni) yozing.")
            return redirect("project_detail", pk=pk)
        if not _qayt_kod_yubor(request, "limit", pk, f"{p.code} — {p.name}", sabab):
            messages.error(request, "Telegramga kod yuborib bo'lmadi — bot bog'lanishini tekshiring.")
            return redirect("project_detail", pk=pk)
        return render(request, "projects/qaytarish_kod.html", {
            "sarlavha": "Umumiy limitni PTOga qaytarish",
            "obj_nomi": f"{p.code} — {p.name}", "sabab": sabab,
            "action_url": reverse("limit_return_pto", args=[pk]),
            "bekor_url": reverse("project_detail", args=[pk]),
            "bosqich": "kod",
        })

    # 2-bosqich: kod tekshiriladi
    sabab, xato = _qayt_kod_tekshir(request, "limit", pk, kod)
    if xato:
        messages.error(request, xato)
        if "qaytadan" in xato:
            return redirect("project_detail", pk=pk)
        return render(request, "projects/qaytarish_kod.html", {
            "sarlavha": "Umumiy limitni PTOga qaytarish",
            "obj_nomi": f"{p.code} — {p.name}", "sabab": request.POST.get("sabab", ""),
            "action_url": reverse("limit_return_pto", args=[pk]),
            "bekor_url": reverse("project_detail", args=[pk]),
            "bosqich": "kod",
        })
    with transaction.atomic():
        req = LimitChangeRequest.objects.create(
            project=p,
            old_material=p.limit_material, old_labor=p.limit_labor,
            old_machinery=p.limit_machinery, old_other=p.limit_other,
            new_material=p.limit_material, new_labor=p.limit_labor,
            new_machinery=p.limit_machinery, new_other=p.limit_other,
            reason=(f"ADMIN QAYTARDI (kamchilik): {sabab}")[:500],
            requested_by=request.user, status="pto2",
        )
        LimitChangeItem.objects.bulk_create([
            LimitChangeItem(request=req, kind=it.kind, name=it.name, unit=it.unit,
                            quantity=it.quantity, unit_price=it.unit_price,
                            note=it.note, bolim=it.bolim, masul=it.masul)
            for it in items
        ])
    messages.success(request,
        "Limit PTOga qaytarildi — PTO «Tasdiqlar»da tahrirlab, direktorga qayta yuboradi.")
    return redirect("project_detail", pk=pk)


@login_required
def weekly_return_pto(request, pk):
    """ADMIN tasdiqlangan HAFTALIK so'rovni PTOga qaytaradi — xuddi limit kabi:
    sabab + Telegram tasdiqlash kodi. So'rov qoralamaga qaytadi, limit tiklanadi."""
    req = get_object_or_404(WeeklyRequest.objects.select_related("project"), pk=pk)
    p = _firma_yoki_403(request, req.project)
    if not is_admin(request.user):
        raise PermissionDenied("Haftalikni faqat admin qaytaradi.")
    if req.status != WeeklyRequest.Status.APPROVED:
        messages.error(request, "Faqat tasdiqlangan haftalik so'rov qaytariladi.")
        return redirect("project_detail", pk=p.pk)
    nomi = f"{p.code} — {req.week_start:%d.%m.%Y} haftalik"

    if request.method != "POST":
        return render(request, "projects/qaytarish_kod.html", {
            "sarlavha": "Haftalik so'rovni PTOga qaytarish",
            "obj_nomi": nomi, "sabab": "",
            "action_url": reverse("weekly_return_pto", args=[pk]),
            "bekor_url": reverse("project_detail", args=[p.pk]),
            "bosqich": "sabab",
        })

    kod = (request.POST.get("kod") or "").strip()
    if not kod:
        sabab = (request.POST.get("sabab") or "").strip()
        if not sabab:
            messages.error(request, "Qaytarish sababini (kamchilikni) yozing.")
            return redirect("weekly_return_pto", pk=pk)
        if not _qayt_kod_yubor(request, "haftalik", pk, nomi, sabab):
            messages.error(request, "Telegramga kod yuborib bo'lmadi — bot bog'lanishini tekshiring.")
            return redirect("project_detail", pk=p.pk)
        return render(request, "projects/qaytarish_kod.html", {
            "sarlavha": "Haftalik so'rovni PTOga qaytarish",
            "obj_nomi": nomi, "sabab": sabab,
            "action_url": reverse("weekly_return_pto", args=[pk]),
            "bekor_url": reverse("project_detail", args=[p.pk]),
            "bosqich": "kod",
        })

    sabab, xato = _qayt_kod_tekshir(request, "haftalik", pk, kod)
    if xato:
        messages.error(request, xato)
        if "qaytadan" in xato:
            return redirect("project_detail", pk=p.pk)
        return render(request, "projects/qaytarish_kod.html", {
            "sarlavha": "Haftalik so'rovni PTOga qaytarish",
            "obj_nomi": nomi, "sabab": request.POST.get("sabab", ""),
            "action_url": reverse("weekly_return_pto", args=[pk]),
            "bekor_url": reverse("project_detail", args=[p.pk]),
            "bosqich": "kod",
        })
    req.status = WeeklyRequest.Status.DRAFT
    req.approved_by = None
    req.approved_at = None
    req.reject_note = (f"ADMIN QAYTARDI (kamchilik): {sabab}")[:500]
    req.pto_notified = True
    req.save(update_fields=["status", "approved_by", "approved_at", "reject_note", "pto_notified"])
    messages.success(request,
        "Haftalik so'rov PTOga qaytarildi (qoralama) — limit tiklandi, PTO tahrirlab qayta yuboradi.")
    return redirect("project_detail", pk=p.pk)


def _tasdiqlar_data(status="dir", user=None):
    """Tasdiqlash markazi ma'lumoti: limit o'zgarishlari + haftalik so'rovlar.
    status='dir' — direktor navbati; status='adm' (limit) / 'submitted' (haftalik) — admin navbati.
    user berilsa — faqat o'sha foydalanuvchi firmasidagi so'rovlar (admin=hammasi)."""
    wk_status = "submitted" if status == "adm" else status
    _proj_qs = visible_projects(user) if user is not None else None
    lreqs = LimitChangeRequest.objects.filter(status=status)
    if _proj_qs is not None:
        lreqs = lreqs.filter(project__in=_proj_qs)
    lreqs = (lreqs.select_related("project", "requested_by", "director_by")
             .prefetch_related("proposed_items").order_by("-created_at"))
    LI_CLS = {"material": "mat", "labor": "lab", "machinery": "mach", "other": "oth"}
    lim_list = []
    for r in lreqs:
        # Joriy tarkib bilan solishtirish — tasdiqlashdagi juftlash bilan BIR XIL tartibda,
        # shunda «yangi» / «o'zgargan» belgilari haqiqatga mos bo'ladi
        from collections import defaultdict, deque
        joriy = defaultdict(deque)
        for li in r.project.limit_items.all().order_by("id"):
            joriy[(li.kind, (li.name or "").strip().lower())].append(li)
        items = []
        for it in r.proposed_items.all():
            kalit = (it.kind, (it.name or "").strip().lower())
            navbat = joriy.get(kalit)
            eski = navbat.popleft() if navbat else None
            if eski is None:
                holat, eski_str = "yangi", ""
            elif (eski.quantity != it.quantity or eski.unit_price != it.unit_price
                  or eski.unit != it.unit or eski.note != it.note):
                holat = "ozgargan"
                eski_str = f"Oldin: {_qty(eski.quantity)} × {_money(eski.unit_price)} = {_money(eski.total)}"
            else:
                holat, eski_str = "", ""
            items.append({
                "id": it.id,
                "kind_disp": it.get_kind_display(), "cls": LI_CLS.get(it.kind, "mat"),
                "name": it.name, "unit": it.unit or "—",
                "qty_str": _qty(it.quantity), "price_str": _money(it.unit_price),
                "sum_str": _money(it.total), "izoh": it.note,
                "bolim": it.bolim, "masul": it.masul,
                "holat": holat, "eski_str": eski_str,
                "sana": it.created_at or r.created_at,
            })
        lim_list.append({
            "obj": r,
            "old_str": _money(r.old_total), "new_str": _money(r.new_total),
            "cats": [
                ("Material", _money(r.old_material), _money(r.new_material)),
                ("Ish haqi", _money(r.old_labor), _money(r.new_labor)),
                ("Mashina chasti", _money(r.old_machinery), _money(r.new_machinery)),
                ("Ko'zda tutilmagan", _money(r.old_other), _money(r.new_other)),
            ],
            "items": items,
        })
    wreqs = WeeklyRequest.objects.filter(status=wk_status)
    if _proj_qs is not None:
        wreqs = wreqs.filter(project__in=_proj_qs)
    wreqs = (wreqs.select_related("project", "created_by", "director_by")
             .prefetch_related("items").order_by("-week_start"))
    wk_list = []
    for w in wreqs:
        # Qatorlar — umumiy limit tarkibi bilan solishtiriladi.
        # «YANGI» = umumiy limitda yo'q (tasdiqlangach avtomatik qo'shiladi),
        # «BOSHQA NARX» = limitdagi narxdan farq qiladi.
        limit_map = {}
        for li in w.project.limit_items.all():
            limit_map.setdefault((li.kind, (li.name or "").strip().lower()), li)
        witems = []
        for it in w.items.all():
            li = limit_map.get((it.kind, (it.name or "").strip().lower()))
            if li is None:
                holat, eski_str = "yangi", "Umumiy limitda yo'q — tasdiqlangach ro'yxatga qo'shiladi"
            elif li.unit_price != it.unit_price:
                holat = "ozgargan"
                eski_str = f"Umumiy limitdagi narx: {_money(li.unit_price)}"
            else:
                holat, eski_str = "", ""
            witems.append({
                "id": it.id,
                "kind_disp": it.get_kind_display(), "cls": LI_CLS.get(it.kind, "mat"),
                "name": it.name, "unit": it.unit or "—",
                "qty_str": _qty(it.quantity), "price_str": _money(it.unit_price),
                "sum_str": _money(it.total), "izoh": it.note,
                "holat": holat, "eski_str": eski_str,
                "sana": it.created_at or w.created_at,
            })
        wk_list.append({
            "obj": w,
            "jami_str": _money(sum((it.total for it in w.items.all()), Decimal("0.00"))),
            "soni": len(w.items.all()),
            "items": witems,
            # Admin tasdiqlashdan OLDIN limitdan oshishini ko'rsin
            "oshgan": _limit_oshish(w.project, _items_by_kind(w.items.all()),
                                    exclude_request_id=w.id),
        })
    return lim_list, wk_list


@login_required
def tasdiqlar(request):
    """Asosiy admin uchun tasdiqlash markazi (endi asosan «Limitlar» tab ichida)."""
    if not is_admin(request.user):
        raise PermissionDenied("Faqat asosiy admin.")
    lim_list, wk_list = _tasdiqlar_data()
    return render(request, "projects/tasdiqlar.html", {"lim_list": lim_list, "wk_list": wk_list})


@login_required
def limit_request_action(request, pk):
    """Limit o'zgartirish so'rovini tasdiqlash / rad etish (admin)."""
    req = get_object_or_404(LimitChangeRequest, pk=pk)
    _firma_yoki_403(request, req.project)
    # «ack» — so'rov EGASI xabarni yopadi (admin bo'lishi shart emas)
    if request.method == "POST" and request.POST.get("action") == "ack":
        if req.requested_by_id == request.user.id:
            req.pto_notified = True
            req.save(update_fields=["pto_notified"])
        return redirect("project_detail", pk=req.project_id)
    from .roles import is_asosiy_admin, is_director, is_pto, is_snab
    S = LimitChangeRequest.Status
    if not (is_director(request.user) or is_admin(request.user)
            or is_snab(request.user) or is_pto(request.user)):
        raise PermissionDenied("Faqat snabjeniye, PTO, direktor yoki admin.")
    if request.method == "POST":
        a = request.POST.get("action")
        if a == "delitem":
            # Bitta taklif qatorini o'chirish — so'rov summalari qayta hisoblanadi.
            # Har kim FAQAT O'Z bosqichida o'chiradi: snab->snab, pto->pto2,
            # direktor->dir, admin->adm.
            _stage_ok = (req.status == S.ADM and is_admin(request.user)) or \
                        (req.status == S.DIR and is_director(request.user)) or \
                        (req.status == S.SNAB and is_snab(request.user)) or \
                        (req.status == S.PTO2 and is_pto(request.user))
            it = req.proposed_items.filter(id=request.POST.get("item_id")).first()
            if req.status not in (S.SNAB, S.PTO2, S.DIR, S.ADM):
                messages.error(request, "Bu so'rov allaqachon ko'rib chiqilgan.")
            elif not _stage_ok:
                messages.error(request, "Bu so'rov sizning bosqichingizda emas.")
            elif it is None:
                messages.error(request, "Qator topilmadi.")
            elif req.proposed_items.count() <= 1:
                messages.error(request, "Oxirgi qatorni o'chirib bo'lmaydi — o'rniga so'rovni «Rad etish» qiling.")
            else:
                nom = it.name
                with transaction.atomic():
                    it.delete()
                    sums = {k: Decimal("0") for k in KINDS}
                    for x in req.proposed_items.all():
                        sums[x.kind] += x.total
                    req.new_material = sums["material"]
                    req.new_labor = sums["labor"]
                    req.new_machinery = sums["machinery"]
                    req.new_other = sums["other"]
                    req.save(update_fields=["new_material", "new_labor", "new_machinery", "new_other"])
                messages.info(request, f"«{nom}» qatori o'chirildi. Yangi umumiy: {_money(req.new_total)}.")
            return redirect(reverse("dashboard") + "?tab=tasdiqlar")
        from django.utils import timezone
        # SNABJENIYE bosqichi: snab → pto2 (narxlab PTOga qaytaradi)
        if a == "snab_return":
            if not (is_snab(request.user) or is_admin(request.user)):
                raise PermissionDenied("Faqat snabjeniye.")
            if req.status != S.SNAB:
                messages.error(request, "Bu so'rov snabjeniye bosqichida emas.")
            else:
                req.status = S.PTO2
                req.snab_by = request.user
                req.snab_at = timezone.now()
                req.save(update_fields=["status", "snab_by", "snab_at"])
                messages.success(request, f"«{req.project.name}» — snabjeniye ko'rdi, PTO xulosasiga qaytdi.")
            return redirect(reverse("dashboard") + "?tab=tasdiqlar")
        # PTO XULOSASI bosqichi: pto2 → dir (yakuniy xulosa bilan direktorga)
        if a == "pto_send_dir":
            if not (is_pto(request.user) or is_admin(request.user)):
                raise PermissionDenied("Faqat PTO.")
            if req.status != S.PTO2:
                messages.error(request, "Bu so'rov PTO xulosasi bosqichida emas.")
            else:
                req.status = S.DIR
                req.save(update_fields=["status"])
                messages.success(request, f"«{req.project.name}» — PTO xulosasi bilan direktor tasdig'iga yuborildi.")
            return redirect(reverse("dashboard") + "?tab=tasdiqlar")
        # DIREKTOR bosqichi: dir → adm (yoki rad)
        if a == "dir_approve":
            # Haqiqiy direktor YOKI ASOSIY admin tasdiqlaydi (admin1/2 emas)
            if request.user.is_superuser and not is_asosiy_admin(request.user):
                messages.error(request, "Direktor bosqichini HAQIQIY direktor yoki "
                                        "ASOSIY admin tasdiqlaydi.")
                return redirect(reverse("dashboard") + "?tab=tasdiqlar")
            if not (is_director(request.user) or is_asosiy_admin(request.user)):
                raise PermissionDenied("Faqat direktor.")
            if req.status != S.DIR:
                messages.error(request, "Bu so'rov direktor bosqichida emas.")
            else:
                req.status = S.ADM
                req.director_by = request.user
                req.director_at = timezone.now()
                req.save(update_fields=["status", "director_by", "director_at"])
                messages.success(request, f"«{req.project.name}» — direktor tasdiqladi, admin tasdig'iga o'tdi.")
        elif a == "dir_reject":
            if request.user.is_superuser and not is_asosiy_admin(request.user):
                messages.error(request, "Direktor bosqichini HAQIQIY direktor yoki ASOSIY admin ko'radi.")
                return redirect(reverse("dashboard") + "?tab=tasdiqlar")
            if not (is_director(request.user) or is_asosiy_admin(request.user)):
                raise PermissionDenied("Faqat direktor.")
            izoh = (request.POST.get("decision_note") or "").strip()
            if req.status != S.DIR:
                messages.error(request, "Bu so'rov direktor bosqichida emas.")
            elif not izoh:
                messages.error(request, "Rad etish sababini yozing.")
            else:
                req.reject(request.user, izoh)
                messages.info(request, f"«{req.project.name}» — direktor rad etdi. Sabab: {izoh}")
        # ADMIN bosqichi: adm → approved (yoki rad)
        elif a == "approve":
            if not is_admin(request.user):
                raise PermissionDenied("Faqat asosiy admin.")
            if req.status != S.ADM:
                messages.error(request, "Bu so'rov admin bosqichida emas (avval direktor tasdiqlaydi).")
            else:
                req.approve(request.user)
                messages.success(request, f"«{req.project.name}» limiti yangilandi: {_money(req.new_total)}.")
        elif a == "reject":
            if not is_admin(request.user):
                raise PermissionDenied("Faqat asosiy admin.")
            izoh = (request.POST.get("decision_note") or "").strip()
            if req.status != S.ADM:
                messages.error(request, "Bu so'rov admin bosqichida emas.")
            elif not izoh:
                messages.error(request, "Rad etish sababini yozing.")
            else:
                req.reject(request.user, izoh)
                messages.info(request, f"«{req.project.name}» limit so'rovi rad etildi. Sabab: {izoh}")
    return redirect(reverse("dashboard") + "?tab=tasdiqlar")


@login_required
def limit_request_edit(request, pk):
    """Kutilayotgan limit so'rovini tahrirlash — HAR KIM O'Z bosqichida:
    snabjeniye -> snab (narxlaydi), PTO -> pto2 (yakuniy xulosa), admin -> istalgan bosqich."""
    from .roles import is_pto as _is_pto, is_snab as _is_snab
    req = get_object_or_404(
        LimitChangeRequest.objects.select_related("project", "requested_by"), pk=pk)
    S = LimitChangeRequest.Status
    if req.status not in (S.SNAB, S.PTO2, S.DIR, S.ADM):
        messages.error(request, "Bu so'rov allaqachon ko'rib chiqilgan — tahrirlab bo'lmaydi.")
        return redirect(reverse("dashboard") + "?tab=tasdiqlar")
    _ruxsat = is_admin(request.user) or \
        (req.status == S.SNAB and _is_snab(request.user)) or \
        (req.status == S.PTO2 and _is_pto(request.user))
    if not _ruxsat:
        raise PermissionDenied("Bu bosqichda so'rovni siz tahrirlay olmaysiz.")
    _firma_yoki_403(request, req.project)
    p = req.project

    if request.method == "POST":
        kinds = request.POST.getlist("kind")
        names = request.POST.getlist("name")
        units = request.POST.getlist("unit")
        qtys = request.POST.getlist("quantity")
        prices = request.POST.getlist("unit_price")
        inotes = request.POST.getlist("item_note")
        bolims = request.POST.getlist("item_bolim")
        masuls = request.POST.getlist("item_masul")
        valid = set(KINDS)
        items = []
        sums = {k: Decimal("0") for k in KINDS}
        for i in range(len(names)):
            nm = (names[i] or "").strip()
            if not nm:
                continue
            kind = kinds[i] if i < len(kinds) and kinds[i] in valid else "material"
            q = _to_dec(qtys[i] if i < len(qtys) else "0") or Decimal("0")
            pr = _to_dec(prices[i] if i < len(prices) else "0") or Decimal("0")
            if q < 0 or pr < 0:
                continue
            items.append(LimitChangeItem(
                request=req, kind=kind, name=nm,
                unit=(units[i] if i < len(units) else "").strip(),
                quantity=q, unit_price=pr,
                note=" ".join((inotes[i] if i < len(inotes) else "").split())[:500],
                bolim=" ".join((bolims[i] if i < len(bolims) else "").split())[:200],
                masul=" ".join((masuls[i] if i < len(masuls) else "").split())[:120],
            ))
            sums[kind] += (q * pr).quantize(Decimal("0.01"))
        if not items:
            messages.error(request, "Kamida bitta qator kiriting.")
            return redirect("limit_request_edit", pk=pk)
        with transaction.atomic():
            req.proposed_items.all().delete()
            LimitChangeItem.objects.bulk_create(items)
            req.new_material = sums["material"]
            req.new_labor = sums["labor"]
            req.new_machinery = sums["machinery"]
            req.new_other = sums["other"]
            req.save(update_fields=["new_material", "new_labor", "new_machinery", "new_other"])
        messages.success(
            request,
            f"So'rov tahrirlandi (yangi umumiy: {_money(req.new_total)}). "
            "Endi tasdiqlash yoki rad etish mumkin.",
        )
        return redirect(reverse("dashboard") + "?tab=tasdiqlar")

    # GET — forma (taklif qatorlari bilan)
    items = [{
        "kind": it.kind, "name": it.name, "unit": it.unit,
        "quantity": it.quantity, "unit_price": it.unit_price, "note": it.note,
        "bolim": it.bolim, "masul": it.masul,
    } for it in req.proposed_items.all()]

    from ombor.models import Material
    unit_map = {}
    for m in Material.objects.all():
        k = m.name.strip().lower()
        if k and m.unit and k not in unit_map:
            unit_map[k] = m.unit
    mat_names = sorted({m.name for m in Material.objects.all()})

    return render(request, "projects/limit_request_edit.html", {
        "req": req, "p": p, "items": items,
        "unit_map": unit_map, "mat_names": mat_names,
    })


@login_required
def limit_bulk(request):
    """Barcha obyektlar limitini bitta ekranda kiritish (firma bo'yicha guruh)."""
    if not is_pto(request.user):
        raise PermissionDenied("Limitni faqat PTO kiritadi.")
    admin = request.user.is_superuser

    from .models import Firma
    firma_id = request.GET.get("firma") or ""

    if request.method == "POST":
        n = 0
        sorov = 0
        for p in visible_projects(request.user):
            editable = admin or p.budget_total <= 0
            if not editable:
                continue
            mat = _to_dec(request.POST.get(f"m_{p.id}"))
            lab = _to_dec(request.POST.get(f"l_{p.id}"))
            mach = _to_dec(request.POST.get(f"k_{p.id}"))
            oth = _to_dec(request.POST.get(f"o_{p.id}")) or Decimal("0")
            if None in (mat, lab, mach) or mat < 0 or lab < 0 or mach < 0 or oth < 0:
                continue
            if (mat, lab, mach, oth) == (p.limit_material, p.limit_labor, p.limit_machinery, p.limit_other):
                continue
            if admin:
                p.limit_material, p.limit_labor, p.limit_machinery, p.limit_other = mat, lab, mach, oth
                p.save(update_fields=["limit_material", "limit_labor", "limit_machinery", "limit_other"])
                n += 1
            else:
                # PTO — limit tasdiqlash zanjiriga so'rov bo'lib boradi
                # (snabjeniye bo'lsa avval unga, keyin PTO -> direktor -> admin)
                if p.limit_requests.filter(status__in=LIM_JARAYON).exists():
                    continue
                LimitChangeRequest.objects.create(
                    project=p,
                    old_material=p.limit_material, old_labor=p.limit_labor,
                    old_machinery=p.limit_machinery, old_other=p.limit_other,
                    new_material=mat, new_labor=lab, new_machinery=mach, new_other=oth,
                    reason="Limit kiritish (jadval)", requested_by=request.user,
                    status=_limit_boshlangich(p),
                )
                sorov += 1
        if admin:
            messages.success(request, f"{n} ta obyekt limiti saqlandi.")
        else:
            messages.success(request, f"{sorov} ta obyekt limiti tasdiqlash zanjiriga yuborildi.")
        url = reverse("dashboard")
        q = "?tab=kiritish" + (f"&firma={firma_id}" if firma_id else "")
        return redirect(url + q)

    # xulosa: nechta obyektga limit belgilangan
    all_projects = list(visible_projects(request.user))
    jami_obj = len(all_projects)
    limitli_obj = sum(1 for p in all_projects if p.budget_total > 0)

    firmalar_qs = visible_firmas(request.user).order_by("name")
    if firma_id:
        firmalar_qs = firmalar_qs.filter(id=firma_id)
    firms = []
    for f in firmalar_qs:
        rows = [{"p": p, "editable": admin or p.budget_total <= 0}
                for p in visible_projects(request.user).filter(firma=f).order_by("code")]
        if rows:
            firms.append({"firma": f, "rows": rows})
    no_firm = []
    if not firma_id and admin:
        no_firm = [{"p": p, "editable": True}
                   for p in Project.objects.filter(firma__isnull=True).order_by("code")]
    return render(request, "projects/limit_bulk.html", {
        "firms": firms, "no_firm": no_firm, "admin": admin,
        "firmalar": visible_firmas(request.user).order_by("name"),
        "sel_firma": firma_id,
        "jami_obj": jami_obj, "limitli_obj": limitli_obj,
    })


@login_required
def weekly_add(request, pk):
    """PTO yangi haftalik so'rov (limit) qo'shadi — qoralama sifatida."""
    p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))
    if not is_pto(request.user):
        raise PermissionDenied("Haftalik so'rovni faqat PTO qo'shadi.")
    if request.method != "POST":
        return redirect("project_detail", pk=pk)

    try:
        ws = datetime.date.fromisoformat(request.POST.get("week_start", ""))
        we = datetime.date.fromisoformat(request.POST.get("week_end", ""))
    except ValueError:
        messages.error(request, "Hafta boshi va oxiri sanasini to'g'ri kiriting.")
        return redirect("project_detail", pk=pk)

    xato = _hafta_sana_xatosi(p, ws, we)
    if xato:
        messages.error(request, xato)
        return redirect("project_detail", pk=pk)

    req = WeeklyRequest.objects.create(
        project=p, week_start=ws, week_end=we,
        number=(request.POST.get("number") or "").strip(),
        note=(request.POST.get("note") or "").strip(),
        status="draft", created_by=request.user,
    )
    kinds = request.POST.getlist("kind")
    names = request.POST.getlist("name")
    units = request.POST.getlist("unit")
    qtys = request.POST.getlist("quantity")
    prices = request.POST.getlist("unit_price")
    # Qator izohi — «note» so'rovning umumiy izohi uchun band, shuning uchun «item_note»
    inotes = request.POST.getlist("item_note")
    valid_kinds = set(KINDS)
    n = 0
    tashlangan = []   # nomi bor, lekin miqdor/narxi to'liq emas — indamay yo'qotmaymiz
    for i in range(len(names)):
        name = (names[i] or "").strip()
        q = _to_dec(qtys[i] if i < len(qtys) else None)
        pr = _to_dec(prices[i] if i < len(prices) else None)
        if not name:
            continue
        if q is None or pr is None:
            yetishmaydi = []
            if q is None:
                yetishmaydi.append("miqdor")
            if pr is None:
                yetishmaydi.append("narx")
            tashlangan.append(f"«{name}» ({' va '.join(yetishmaydi)} kiritilmagan)")
            continue
        kind = kinds[i] if i < len(kinds) else "material"
        if kind not in valid_kinds:
            kind = "material"
        WeeklyRequestItem.objects.create(
            request=req, kind=kind,
            name=name, unit=(units[i] if i < len(units) else "").strip(),
            quantity=q, unit_price=pr,
            note=(inotes[i] if i < len(inotes) else "").strip()[:500],
        )
        n += 1
    if n == 0:
        req.delete()
        if tashlangan:
            messages.error(request, "So'rov saqlanmadi — hech bir qator to'liq emas: " + "; ".join(tashlangan))
        else:
            messages.error(request, "Kamida bitta to'liq qator (nomi, objём, narx) kiriting.")
        return redirect("project_detail", pk=pk)

    if tashlangan:
        messages.warning(
            request,
            f"{len(tashlangan)} ta qator saqlanmadi (to'liq emas): " + "; ".join(tashlangan),
        )

    # BIR BOSQICHLI oqim — umumiy limitdagidek:
    #   admin  -> darrov qo'llanadi (u tasdiqlovchi)
    #   PTO    -> to'g'ridan-to'g'ri admin tasdig'iga ketadi (qoralama bosqichi yo'q)
    # Limitdan oshsa admin ham darrov tasdiqlay olmaydi — «Tasdiqlar»da ataylab tasdiqlaydi.
    oshgan = _limit_oshish(p, _items_by_kind(req.items.all()), exclude_request_id=req.id)
    if is_admin(request.user) and not oshgan:
        from django.utils import timezone as _tz
        req.status = WeeklyRequest.Status.APPROVED
        req.approved_by = request.user
        req.approved_at = _tz.now()
        req.save(update_fields=["status", "approved_by", "approved_at"])
        messages.success(request, f"Haftalik so'rov saqlandi va tasdiqlandi ({n} qator) — limitdan ayirildi.")
    else:
        req.status = WeeklyRequest.Status.DIR
        req.save(update_fields=["status"])
        messages.success(request, f"Haftalik so'rov direktor tasdig'iga yuborildi ({n} qator).")
        if oshgan:
            messages.error(
                request,
                "DIQQAT — so'rov limitdan oshadi: " + "; ".join(oshgan)
                + ". Admin «Tasdiqlar» bo'limida buni alohida tasdiqlashi kerak.",
            )
    return redirect("project_detail", pk=pk)


@login_required
def weekly_action(request, pk):
    """Haftalik so'rov oqimi: PTO yuboradi → Direktor tasdiqlaydi → Admin tasdiqlaydi."""
    from .roles import is_admin, is_director
    req = get_object_or_404(WeeklyRequest, pk=pk)
    _firma_yoki_403(request, req.project)
    WS = WeeklyRequest.Status
    proj_id = req.project_id
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "ack":
            # So'rov egasi «Tasdiqlandi» xabarini ko'rdi — boshqa ko'rsatilmaydi
            if req.created_by_id == request.user.id:
                req.pto_notified = True
                req.save(update_fields=["pto_notified"])
            return redirect("project_detail", pk=proj_id)
        if action == "delitem":
            # Bitta qatorni so'rovdan o'chirish — har kim FAQAT O'Z bosqichida:
            # direktor -> dir, admin -> submitted (aks holda direktor admin
            # navbatidagi tarkibni o'zgartirib qo'yadi).
            if not (is_director(request.user) or is_admin(request.user)):
                raise PermissionDenied("Qatorni faqat direktor yoki admin o'chiradi.")
            _stage_ok = (req.status == WS.SUBMITTED and is_admin(request.user)) or \
                        (req.status == WS.DIR and is_director(request.user))
            it = req.items.filter(id=request.POST.get("item_id")).first()
            if req.status == WeeklyRequest.Status.APPROVED:
                messages.error(request, "Tasdiqlangan so'rov qatorini o'chirib bo'lmaydi.")
            elif not _stage_ok:
                messages.error(request, "Bu so'rov sizning bosqichingizda emas.")
            elif it is None:
                messages.error(request, "Qator topilmadi.")
            elif req.items.count() <= 1:
                messages.error(request, "Oxirgi qatorni o'chirib bo'lmaydi — o'rniga so'rovni qaytaring.")
            else:
                nom = it.name
                it.delete()
                messages.info(request, f"«{nom}» qatori so'rovdan o'chirildi.")
            if request.POST.get("next") == "tasdiqlar":
                return redirect(reverse("dashboard") + "?tab=tasdiqlar")
            return redirect("project_detail", pk=proj_id)
        if action == "submit":
            # PTO qoralamani DIREKTORga yuboradi — shundan keyingina tasdiqlar zanjiriga tushadi
            if not is_pto(request.user):
                raise PermissionDenied("Faqat PTO yuboradi.")
            if req.status != WS.DRAFT:
                messages.error(request, "Faqat qoralama holatidagi so'rov yuboriladi.")
            else:
                req.status = WS.DIR
                req.save(update_fields=["status"])
                messages.success(request, "So'rov direktor tasdig'iga yuborildi.")
        elif action == "recall":
            # PTO yuborilganini qaytarib oladi (direktor hali ko'rmagan bo'lsa)
            if not is_pto(request.user):
                raise PermissionDenied("Faqat PTO.")
            if req.status != WS.DIR:
                messages.error(request, "Faqat direktor tasdig'idagi so'rov qaytarib olinadi.")
            else:
                req.status = WS.DRAFT
                req.save(update_fields=["status"])
                messages.info(request, "So'rov qoralamaga qaytarib olindi.")
        elif action == "dir_approve":
            # DIREKTOR (yoki ASOSIY admin) tasdiqlaydi: dir → submitted
            from .roles import is_asosiy_admin as _asos
            if request.user.is_superuser and not _asos(request.user):
                messages.error(request, "Direktor bosqichini HAQIQIY direktor yoki "
                                        "ASOSIY admin tasdiqlaydi.")
                return redirect(reverse("dashboard") + "?tab=tasdiqlar")
            if not (is_director(request.user) or _asos(request.user)):
                raise PermissionDenied("Faqat direktor.")
            if req.status != WS.DIR:
                messages.error(request, "Bu so'rov direktor bosqichida emas.")
            else:
                from django.utils import timezone
                req.status = WS.SUBMITTED
                req.director_by = request.user
                req.director_at = timezone.now()
                req.save(update_fields=["status", "director_by", "director_at"])
                messages.success(request, "Direktor tasdiqladi — admin tasdig'iga o'tdi.")
        elif action == "dir_reject":
            from .roles import is_asosiy_admin as _asos2
            if request.user.is_superuser and not _asos2(request.user):
                messages.error(request, "Direktor bosqichini HAQIQIY direktor yoki ASOSIY admin ko'radi.")
                return redirect(reverse("dashboard") + "?tab=tasdiqlar")
            if not (is_director(request.user) or _asos2(request.user)):
                raise PermissionDenied("Faqat direktor.")
            izoh = (request.POST.get("reject_note") or request.POST.get("decision_note") or "").strip()
            if req.status != WS.DIR:
                messages.error(request, "Bu so'rov direktor bosqichida emas.")
            elif not izoh:
                messages.error(request, "Rad etish sababini yozing.")
            else:
                req.status = WS.DRAFT
                req.reject_note = izoh[:500]
                req.save(update_fields=["status", "reject_note"])
                messages.info(request, f"Direktor rad etdi — PTO qoralamasiga qaytdi. Sabab: {izoh}")
        elif action in ("approve", "draft"):
            if not is_admin(request.user):
                raise PermissionDenied("Haftalik limitni tasdiqlash/bekor qilish — faqat asosiy admin.")
            if action == "approve":
                from django.utils import timezone
                # Faqat DIREKTORdan o'tgan (submitted) so'rov admin tomonidan tasdiqlanadi
                if req.status != WS.SUBMITTED:
                    xm = ("Bu so'rov allaqachon tasdiqlangan." if req.status == WS.APPROVED
                          else "Avval direktor tasdiqlashi kerak.")
                    messages.error(request, xm)
                    if request.POST.get("next") == "tasdiqlar":
                        return redirect(reverse("dashboard") + "?tab=tasdiqlar")
                    return redirect("project_detail", pk=proj_id)
                # Tekshiruv + saqlash BITTA tranzaksiyada — ikki parallel tasdiqlash
                # ikkalasi ham limitga sig'ib qolmasin (poyga holati)
                with transaction.atomic():
                    req = (WeeklyRequest.objects.select_for_update()
                           .select_related("project").get(pk=req.pk))
                    oshgan = _limit_oshish(
                        req.project, _items_by_kind(req.items.all()),
                        exclude_request_id=req.id,
                    )
                    if oshgan and request.POST.get("confirm_over") != "1":
                        messages.error(
                            request,
                            "Tasdiqlanmadi — limitdan oshadi: " + "; ".join(oshgan)
                            + ". Baribir tasdiqlash uchun «Limitdan oshsa ham tasdiqlash» tugmasini bosing.",
                        )
                        if request.POST.get("next") == "tasdiqlar":
                            return redirect(reverse("dashboard") + "?tab=tasdiqlar")
                        return redirect("project_detail", pk=proj_id)
                    req.status = "approved"
                    req.approved_by = request.user
                    req.approved_at = timezone.now()
                    req.pto_notified = False   # PTO'ga "Tasdiqlandi" xabari chiqsin
                    req.save(update_fields=["status", "approved_by", "approved_at", "pto_notified"])
                    # Umumiy limitda yo'q materiallar shu yerda ro'yxatga qo'shiladi
                    qoshildi = _limitga_yangi_materiallar(req)
                if qoshildi:
                    messages.info(
                        request,
                        "Umumiy limit ro'yxatiga yangi material qo'shildi (miqdor 0 — "
                        "limit summasi oshmadi, kerak bo'lsa «Umumiy limit ichi»da miqdor belgilang): "
                        + ", ".join(qoshildi),
                    )
                if oshgan:
                    messages.warning(
                        request,
                        "So'rov LIMITDAN OSHIB tasdiqlandi (admin ruxsati bilan): " + "; ".join(oshgan),
                    )
                else:
                    messages.success(request, "So'rov tasdiqlandi — umumiy limitdan ayirildi.")
            else:
                if req.status == WeeklyRequest.Status.APPROVED:
                    # Tasdiqlanganini qaytarish — faqat KOD bilan (weekly_return_pto)
                    messages.error(request,
                        "Tasdiqlangan so'rov faqat tasdiqlash KODI bilan qaytariladi — "
                        "«PTOga qaytarish» tugmasidan foydalaning.")
                    return redirect("project_detail", pk=proj_id)
                req.status = "draft"
                req.approved_by = None
                req.approved_at = None
                req.save(update_fields=["status", "approved_by", "approved_at"])
                messages.info(request, "So'rov qoralamaga qaytarildi — limitga qaytdi.")
        elif action == "delete":
            if not is_pto(request.user):
                raise PermissionDenied("Faqat PTO.")
            # Tasdiqlangan so'rovni HECH KIM (admin ham) to'g'ridan-to'g'ri o'chira olmaydi —
            # aks holda tasdiqlangan sarf izsiz yo'qoladi. Avval qoralamaga qaytarilsin.
            if req.status == WeeklyRequest.Status.APPROVED:
                messages.error(
                    request,
                    "Tasdiqlangan so'rovni o'chirib bo'lmaydi. Avval admin uni "
                    "«Qoralamaga qaytarish» orqali bekor qilsin, so'ng o'chirish mumkin.",
                )
            elif req.status in (WeeklyRequest.Status.DIR, WeeklyRequest.Status.SUBMITTED):
                messages.error(request, "Yuborilgan so'rovni avval «Qaytarib olish» qiling, so'ng o'chirasiz.")
            else:
                req.delete()
                messages.info(request, "So'rov o'chirildi.")
    if request.POST.get("next") == "tasdiqlar":
        return redirect(reverse("dashboard") + "?tab=tasdiqlar")
    return redirect("project_detail", pk=proj_id)


@login_required
def weekly_edit(request, pk):
    """Haftalik so'rovni tahrirlash — FAQAT admin (to'g'ridan-to'g'ri qo'llanadi)."""
    req = get_object_or_404(WeeklyRequest.objects.select_related("project"), pk=pk)
    if not is_admin(request.user):
        raise PermissionDenied("Haftalik so'rovni faqat asosiy admin tahrirlaydi.")
    # Admin har qanday holatda (tasdiqlangan bo'lsa ham) tahrirlay oladi.
    # Sarf jonli hisoblanadi (tasdiqlangan qatorlar yig'indisi) — tahrir darrov aks etadi.
    was_approved = req.status == WeeklyRequest.Status.APPROVED
    p = req.project

    if request.method == "POST":
        try:
            ws = datetime.date.fromisoformat(request.POST.get("week_start", ""))
            we = datetime.date.fromisoformat(request.POST.get("week_end", ""))
        except ValueError:
            messages.error(request, "Hafta boshi va oxiri sanasini to'g'ri kiriting.")
            return redirect("weekly_edit", pk=pk)

        xato = _hafta_sana_xatosi(p, ws, we, exclude_id=req.id)
        if xato:
            messages.error(request, xato)
            return redirect("weekly_edit", pk=pk)

        kinds = request.POST.getlist("kind")
        names = request.POST.getlist("name")
        units = request.POST.getlist("unit")
        qtys = request.POST.getlist("quantity")
        prices = request.POST.getlist("unit_price")
        inotes = request.POST.getlist("item_note")
        valid = set(KINDS)
        new_items = []
        for i in range(len(names)):
            nm = (names[i] or "").strip()
            if not nm:
                continue
            kind = kinds[i] if i < len(kinds) and kinds[i] in valid else "material"
            unit = (units[i] if i < len(units) else "").strip()
            q = _to_dec(qtys[i] if i < len(qtys) else "0") or Decimal("0")
            pr = _to_dec(prices[i] if i < len(prices) else "0") or Decimal("0")
            if q < 0 or pr < 0:
                continue
            new_items.append(WeeklyRequestItem(request=req, kind=kind, name=nm, unit=unit,
                                               quantity=q, unit_price=pr,
                                               note=(inotes[i] if i < len(inotes) else "").strip()[:500]))
        if not new_items:
            messages.error(request, "Kamida bitta qator kiriting.")
            return redirect("weekly_edit", pk=pk)

        req.week_start = ws
        req.week_end = we
        req.number = (request.POST.get("number") or "").strip()
        req.note = (request.POST.get("note") or "").strip()
        with transaction.atomic():
            req.save(update_fields=["week_start", "week_end", "number", "note"])
            req.items.all().delete()
            WeeklyRequestItem.objects.bulk_create(new_items)
            if was_approved:
                # Tasdiqlangan so'rov tahrirlangach yangi materiallarni ham qo'shamiz
                _limitga_yangi_materiallar(req)
        messages.success(request, "Haftalik so'rov tahrirlandi.")
        if was_approved:
            # Tahrir tasdiqlangan sarfni o'zgartirdi — limitdan oshsa admin bilsin
            oshgan = _limit_oshish(p, _items_by_kind(req.items.all()), exclude_request_id=req.id)
            if oshgan:
                messages.warning(request, "DIQQAT — tahrirdan keyin limitdan oshdi: " + "; ".join(oshgan))
        return redirect("project_detail", pk=p.id)

    # GET — tahrir formasi (prefilled)
    items = [{
        "kind": it.kind, "name": it.name, "unit": it.unit,
        "quantity": it.quantity, "unit_price": it.unit_price,
        "note": it.note,
    } for it in req.items.all()]

    # birlik avtomat uchun nom->birlik xaritasi
    from ombor.models import Material
    unit_map = {}
    for m in Material.objects.all():
        k = m.name.strip().lower()
        if k and m.unit and k not in unit_map:
            unit_map[k] = m.unit
    for it in WeeklyRequestItem.objects.filter(kind="material").exclude(unit="").values("name", "unit"):
        k = (it["name"] or "").strip().lower()
        if k and k not in unit_map:
            unit_map[k] = it["unit"]
    mat_names = sorted({m.name for m in Material.objects.all()})

    return render(request, "projects/weekly_edit.html", {
        "req": req, "p": p, "items": items,
        "unit_map": unit_map, "mat_names": mat_names,
    })


def build_hisobotlar_zip(user=None):
    """Barcha hisobotlarni bitta ZIP (bytes) qilib qaytaradi:
       - Umumiy_limitlar.xlsx (barcha obyektlar jadvali)
       - Obyektlar/<kod>_limit.xlsx (har obyekt: limit ichi + haftalik)
       - Ombor_qoldiq.xlsx (joriy qoldiq), agar mavjud bo'lsa.
       View ham, Telegram bot ham shu funksiyani ishlatadi.
       user berilsa — faqat o'sha foydalanuvchi firmasidagi obyektlar (bot=None=hammasi)."""
    import zipfile

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    proj_qs = visible_projects(user) if user is not None else Project.objects.all()

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as z:
        # 1) Umumiy limitlar jadvali
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Limitlar"
        ws.append(["Kod", "Nomi", "Firma", "Material", "Ish haqi", "Mashina chasti",
                   "Ko'zda tutilmagan", "Umumiy limit", "Sarflangan", "Qolgan"])
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="213145")
        for p in proj_qs.select_related("firma").order_by("code"):
            sarf = p.sarflangan()
            lim = p.budget_total or Decimal("0")
            ws.append([p.code, p.name, p.firma.name if p.firma_id else "—",
                       float(p.limit_material or 0), float(p.limit_labor or 0),
                       float(p.limit_machinery or 0), float(p.limit_other or 0),
                       float(lim), float(sarf), float(lim - sarf)])
        ws.freeze_panes = "A2"
        for i, w in enumerate([16, 30, 22, 16, 16, 18, 18, 18, 18, 18], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        b = io.BytesIO(); wb.save(b)
        z.writestr("Umumiy_limitlar.xlsx", b.getvalue())

        # 2) Har obyekt uchun batafsil (faqat limiti yoki haftaligi bo'lganlar)
        for p in proj_qs.order_by("code"):
            if p.budget_total <= 0 and not p.weekly_requests.exists() and not p.limit_items.exists():
                continue
            b = io.BytesIO(); _obj_limit_wb(p).save(b)
            fn = f"Obyektlar/{p.code}_limit.xlsx".replace(" ", "_")
            z.writestr(fn, b.getvalue())

        # 3) Ombor qoldig'i (foydalanuvchi ko'rish doirasi bo'yicha; bot user=None=hammasi)
        try:
            from ombor.views import _ombor_wb
            b = io.BytesIO(); _ombor_wb(user=user).save(b)
            z.writestr("Ombor_qoldiq.xlsx", b.getvalue())
        except Exception:
            pass

    return zbuf.getvalue()


@login_required
def hisobotlar_zip(request):
    """Barcha hisobotlarni bitta ZIP faylga jamlab yuklab olish."""
    from django.utils import timezone as _tz
    resp = HttpResponse(build_hisobotlar_zip(user=request.user), content_type="application/zip")
    stamp = _tz.localtime().strftime("%Y%m%d_%H%M")
    resp["Content-Disposition"] = f'attachment; filename="hisobotlar_{stamp}.zip"'
    return resp


@login_required
def limit_export(request):
    """Obyektlar limitini Excel'ga eksport (qayta import qilsa bo'ladigan format)."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Limitlar"
    headers = ["Kod", "Nomi", "Material limiti", "Ish haqi limiti", "Mashina chasti limiti",
               "Ko'zda tutilmagan limiti", "Umumiy limit", "Sarflangan", "Qolgan"]
    ws.append(headers)
    for p in visible_projects(request.user).order_by("code"):
        sarf = p.sarflangan()
        limit = p.budget_total or Decimal("0")
        ws.append([
            p.code, p.name,
            float(p.limit_material or 0), float(p.limit_labor or 0),
            float(p.limit_machinery or 0), float(p.limit_other or 0),
            float(limit), float(sarf), float(limit - sarf),
        ])
    for i, w in enumerate([16, 30, 18, 18, 20, 20, 18, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="limitlar.xlsx"'
    return resp


def _obj_limit_wb(p):
    """Bitta obyekt uchun limit hisobotini (2 varaqli) openpyxl Workbook qaytaradi.
       1-varaq — Umumiy limit ichi; 2-varaq — Haftalik so'rovlar."""
    from django.utils.timezone import localtime as _tzloc
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    KIND = dict(KIND_NOMI)
    STATUS = {"draft": "Qoralama", "dir": "Direktor tasdig'ida",
              "submitted": "Admin tasdig'ida", "approved": "Tasdiqlangan"}

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="213145")
    title_font = Font(bold=True, size=14, color="0B1C30")
    sub_font = Font(size=10, color="64748B")
    tot_font = Font(bold=True, size=11)
    tot_fill = PatternFill("solid", fgColor="EFF4FF")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '# ##0'
    right = Alignment(horizontal="right")

    def style_head(ws, row):
        for c in ws[row]:
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()

    # ===== 1-VARAQ: Umumiy limit ichi =====
    ws1 = wb.active
    ws1.title = "Umumiy limit"
    ws1["A1"] = f"{p.code} — {p.name}"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Umumiy limit tarkibi · Jami: {float(p.budget_total):,.0f} so'm".replace(",", " ")
    ws1["A2"].font = sub_font
    ws1.append([])
    hrow = 4
    ws1.append(["Bo'lim", "Mas'ul", "Turi", "Nomi", "Izoh", "Birlik", "Miqdor", "Narxi", "Summa", "Qo'shilgan sana"])
    style_head(ws1, hrow)
    for it in p.limit_items.all().order_by("id"):
        ws1.append([
            it.bolim or "", it.masul or "",
            KIND.get(it.kind, it.kind), it.name, it.note or "", it.unit or "",
            float(it.quantity), float(it.unit_price), float(it.total),
            _tzloc(it.created_at).strftime("%d.%m.%Y %H:%M") if it.created_at else "",
        ])
    last = ws1.max_row
    for r in range(hrow + 1, last + 1):
        for col in (7, 8, 9):
            ws1.cell(r, col).number_format = money
            ws1.cell(r, col).alignment = right
        for c in ws1[r]:
            c.border = border
    # kategoriya bo'yicha jami
    ws1.append([])
    d = {k: Decimal("0") for k in KINDS}
    for it in p.limit_items.all():
        d[it.kind] = d.get(it.kind, Decimal("0")) + it.total
    for k in KINDS:
        rr = ws1.max_row + 1
        ws1.cell(rr, 8, KIND[k] + " jami:").font = tot_font
        c = ws1.cell(rr, 9, float(d[k])); c.number_format = money; c.font = tot_font; c.fill = tot_fill
    rr = ws1.max_row + 1
    ws1.cell(rr, 8, "UMUMIY LIMIT:").font = Font(bold=True, size=12)
    c = ws1.cell(rr, 9, float(p.budget_total)); c.number_format = money
    c.font = Font(bold=True, size=12, color="2563EB"); c.fill = tot_fill
    ws1.freeze_panes = "A5"
    for i, w in enumerate([22, 16, 16, 34, 30, 10, 12, 14, 16, 16], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ===== 2-VARAQ: Haftalik so'rovlar =====
    ws2 = wb.create_sheet("Haftalik so'rovlar")
    ws2["A1"] = f"{p.code} — {p.name}"
    ws2["A1"].font = title_font
    ws2["A2"] = "Haftalik so'rovlar (barcha qatorlar bilan)"
    ws2["A2"].font = sub_font
    ws2.append([])
    ws2.append(["Hafta", "Holati", "Hujjat №", "Turi", "Nomi", "Izoh",
                "Birlik", "Miqdor", "Narxi", "Summa", "Kiritdi", "Sana"])
    style_head(ws2, 4)
    reqs = (p.weekly_requests.select_related("created_by")
            .prefetch_related("items").order_by("-week_start", "-id"))
    for w in reqs:
        hafta = f"{w.week_start:%d.%m.%Y} — {w.week_end:%d.%m.%Y}"
        items = list(w.items.all())
        if not items:
            ws2.append([hafta, STATUS.get(w.status, w.status), w.number or "", "", "(qatorsiz)"])
            continue
        for it in items:
            ws2.append([
                hafta, STATUS.get(w.status, w.status), w.number or "",
                KIND.get(it.kind, it.kind), it.name, it.note or "", it.unit or "",
                float(it.quantity), float(it.unit_price), float(it.total),
                w.created_by.username if w.created_by_id else "—",
                _tzloc(it.created_at).strftime("%d.%m.%Y %H:%M") if it.created_at else "",
            ])
    for r in range(5, ws2.max_row + 1):
        for col in (8, 9, 10):
            ws2.cell(r, col).number_format = money
            ws2.cell(r, col).alignment = right
        for c in ws2[r]:
            c.border = border
    ws2.freeze_panes = "A5"
    for i, w in enumerate([22, 18, 12, 15, 30, 26, 10, 12, 14, 16, 14, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    return wb


@login_required
def limit_export_obj(request, pk):
    """Bitta obyekt limit hisoboti — Excel."""
    p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))
    buf = io.BytesIO()
    _obj_limit_wb(p).save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fn = f"limit_{p.code}.xlsx".replace(" ", "_")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp


@login_required
def limit_template(request):
    """Limit import uchun tayyor Excel qolip (joriy obyektlar bilan)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Limit"
    headers = ["Kod", "Nomi", "Material limiti", "Ish haqi limiti", "Mashina chasti limiti",
               "Ko'zda tutilmagan limiti"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")

    for p in visible_projects(request.user).order_by("code"):
        ws.append([p.code, p.name, float(p.limit_material or 0), float(p.limit_labor or 0),
                   float(p.limit_machinery or 0), float(p.limit_other or 0)])
    # bo'sh namuna qatorlari (yangi obyekt qo'shish uchun)
    for _ in range(3):
        ws.append(["", "", "", "", "", ""])

    ws.freeze_panes = "A2"
    for i, w in enumerate([18, 34, 18, 18, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="limit_shablon.xlsx"'
    return resp


@login_required
def grafik_template(request):
    """Haftalik grafik qolipi (bo'sh): Limit (ish haqi + material) + Texnika (alohida)."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="AAB2C0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    HFILL = PatternFill("solid", fgColor="2563EB")
    GFILL = PatternFill("solid", fgColor="EEF2FF")
    DAYFILL = PatternFill("solid", fgColor="1E40AF")
    HFONT = Font(bold=True, color="FFFFFF", size=10)
    BFONT = Font(bold=True, size=10)
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DAYS = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]

    def hcell(ws, r, c, val, fill=HFILL, font=HFONT):
        cell = ws.cell(r, c, val)
        cell.fill = fill
        cell.font = font
        cell.alignment = CEN
        cell.border = BORDER
        return cell

    def merge(ws, r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    wb = openpyxl.Workbook()

    # ===================== 1) LIMIT =====================
    ws = wb.active
    ws.title = "Limit"
    ws["A1"] = "Obyekt:"
    ws["A1"].font = BFONT
    ws["B1"] = "________________________"
    ws["E1"] = "Hafta boshi (Dushanba) sanasi →"
    ws["E1"].font = BFONT
    ws["H1"].number_format = "dd.mm.yyyy"
    ws["H1"].font = BFONT

    # sarlavha: 3 qatorli (super-guruh / guruh / kichik)
    HR1, HR2, HR3 = 3, 4, 5
    fixed = ["№", "Naimenovaniye (ish / material)", "O'lchov birligi", "Kol-vo"]
    for i, t in enumerate(fixed, start=1):
        hcell(ws, HR1, i, t); merge(ws, HR1, i, HR3, i)

    # --- UMUMIY LIMIT (5-9): Ish haqi + Material + Umumiy ---
    hcell(ws, HR1, 5, "UMUMIY LIMIT"); merge(ws, HR1, 5, HR1, 9)
    hcell(ws, HR2, 5, "Ish haqi"); merge(ws, HR2, 5, HR2, 6)
    hcell(ws, HR3, 5, "Narxi"); hcell(ws, HR3, 6, "Summasi")
    hcell(ws, HR2, 7, "Material"); merge(ws, HR2, 7, HR2, 8)
    hcell(ws, HR3, 7, "Narxi"); hcell(ws, HR3, 8, "Summasi")
    hcell(ws, HR2, 9, "Umumiy summa"); merge(ws, HR2, 9, HR3, 9)

    # --- BAJARILGAN / QOLDIQ (10-11) ---
    BQFILL = PatternFill("solid", fgColor="047857")
    hcell(ws, HR1, 10, "BAJARILGAN / QOLDIQ", fill=BQFILL); merge(ws, HR1, 10, HR1, 11)
    hcell(ws, HR2, 10, "Bajarilgan", fill=BQFILL); merge(ws, HR2, 10, HR3, 10)
    hcell(ws, HR2, 11, "Qoldiq", fill=BQFILL); merge(ws, HR2, 11, HR3, 11)

    # --- HAFTALIK LIMIT (kunlik) ---
    col = 12
    day_start = col
    for i in range(7):
        dc = hcell(ws, HR2, col, f'=IF($H$1="","",$H$1+{i})', fill=DAYFILL)
        dc.number_format = "dddd, dd.mm"  # hafta kuni + sana
        merge(ws, HR2, col, HR2, col + 1)
        hcell(ws, HR3, col, "Miqdor", fill=DAYFILL)
        hcell(ws, HR3, col + 1, "Summa", fill=DAYFILL)
        col += 2
    hcell(ws, HR2, col, "Haftalik jami"); merge(ws, HR2, col, HR3, col)
    last_col = col
    hcell(ws, HR1, day_start, "HAFTALIK LIMIT (sanalar avtomatik)", fill=DAYFILL); merge(ws, HR1, day_start, HR1, last_col)

    # JAMI qatori + formulali data qatorlar
    L = get_column_letter
    jrow = HR3 + 1
    data_first, data_last = jrow + 1, jrow + 34
    day_sum_cols = list(range(day_start + 1, last_col, 2))  # kunlik "Summa" ustunlari
    for r in range(data_first, data_last + 1):
        # summasi = narxi * kol-vo (ish haqi, material)
        ws.cell(r, 6, f'=IF(OR($D{r}="",E{r}=""),"",E{r}*$D{r})')
        ws.cell(r, 8, f'=IF(OR($D{r}="",G{r}=""),"",G{r}*$D{r})')
        # Umumiy summa = ish haqi + material
        ws.cell(r, 9, f'=IF(AND(F{r}="",H{r}=""),"",SUM(F{r},H{r}))')
        # Haftalik jami = kunlik summalar yig'indisi
        rng = ",".join(f"{L(c)}{r}" for c in day_sum_cols)
        ws.cell(r, last_col, f'=IF(SUM({rng})=0,"",SUM({rng}))')
        # Bajarilgan = haftalik jami (kunlik bajarilgan ish) ; Qoldiq = umumiy - bajarilgan
        ws.cell(r, 10, f'=IF({L(last_col)}{r}="","",{L(last_col)}{r})')
        ws.cell(r, 11, f'=IF(I{r}="","",I{r}-N(J{r}))')
        for c in range(1, last_col + 1):
            ws.cell(r, c).border = BORDER
    # JAMI SUMMA qatori — ustun bo'yicha yig'indi
    ws.cell(jrow, 2, "JAMI SUMMA:").font = BFONT
    for c in range(1, last_col + 1):
        ws.cell(jrow, c).fill = GFILL
        ws.cell(jrow, c).border = BORDER
    for c in [6, 8, 9, 10, 11] + day_sum_cols + [last_col]:
        cell = ws.cell(jrow, c, f'=SUM({L(c)}{data_first}:{L(c)}{data_last})')
        cell.font = BFONT

    ws.freeze_panes = ws.cell(jrow, 5).coordinate
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 9
    for c in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # ===================== 2) TEXNIKA (mashina chasti) =====================
    wt = wb.create_sheet("Texnika limit")
    wt["A1"] = "Obyekt:"; wt["A1"].font = BFONT
    wt["B1"] = "________________________"
    wt["E1"] = "Hafta boshi (Dushanba) sanasi →"; wt["E1"].font = BFONT
    wt["H1"].number_format = "dd.mm.yyyy"; wt["H1"].font = BFONT
    fixed2 = ["№", "Bajariladigan ish", "Texnika turi", "O'lchov birligi"]
    for i, t in enumerate(fixed2, start=1):
        hcell(wt, 2, i, t); merge(wt, 2, i, 3, i)
    col = 5
    for i in range(7):
        dc = hcell(wt, 2, col, f'=IF($H$1="","",$H$1+{i})', fill=DAYFILL)
        dc.number_format = "dddd, dd.mm"
        merge(wt, 2, col, 2, col + 2)
        hcell(wt, 3, col, "Miqdor", fill=DAYFILL)
        hcell(wt, 3, col + 1, "Narxi", fill=DAYFILL)
        hcell(wt, 3, col + 2, "Summa", fill=DAYFILL)
        col += 3
    hcell(wt, 2, col, "Haftalik summa"); merge(wt, 2, col, 3, col)
    last2 = col
    day_sum_t = list(range(7, last2, 3))  # har kunning "Summa" ustuni
    for r in range(5, 31):
        for sc in day_sum_t:
            m, n = L(sc - 2), L(sc - 1)  # miqdor, narxi
            wt.cell(r, sc, f'=IF(OR({m}{r}="",{n}{r}=""),"",{m}{r}*{n}{r})')
        rng = ",".join(f"{L(c)}{r}" for c in day_sum_t)
        wt.cell(r, last2, f'=IF(SUM({rng})=0,"",SUM({rng}))')
        for c in range(1, last2 + 1):
            wt.cell(r, c).border = BORDER
    wt.cell(4, 2, "JAMI:").font = BFONT
    for c in range(1, last2 + 1):
        wt.cell(4, c).fill = GFILL; wt.cell(4, c).border = BORDER
    for c in day_sum_t + [last2]:
        wt.cell(4, c, f'=SUM({L(c)}5:{L(c)}30)').font = BFONT
    wt.freeze_panes = "E4"
    wt.column_dimensions["A"].width = 5
    wt.column_dimensions["B"].width = 28
    wt.column_dimensions["C"].width = 20
    wt.column_dimensions["D"].width = 12
    for c in range(5, last2 + 1):
        wt.column_dimensions[get_column_letter(c)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="haftalik_grafik_qolip.xlsx"'
    return resp


GRAFIK_N = 15  # Excel namunadagi kunlik ustunlar soni (eksport oynasi)


def _grafik_kunlar(start, bugun, tugash=None):
    """Veb grafik uzunligi (kun) — HAFTALIK.
    Tugash sanasi berilgan bo'lsa: boshlanishdan tugashgacha (to'liq haftalarga
    yaxlitlab). Berilmasa: joriy hafta + kelasi haftagacha o'sib boradi.
    Kamida 7 kun, ko'pi 84 kun (12 hafta)."""
    if not start:
        return GRAFIK_N
    if tugash and tugash >= start:
        kun = (tugash - start).days + 1
        return max(7, min(84, ((kun + 6) // 7) * 7))
    farq = (bugun - start).days
    if farq < 0:
        return GRAFIK_N
    return max(GRAFIK_N, min(84, (farq // 7 + 2) * 7))


@login_required
def grafik_list(request):
    """Grafik ishlar — obyektlar ro'yxati, har biriga bosib grafik ochiladi."""
    from django.db.models import Count
    firma_id = request.GET.get("firma") or ""
    projs = (visible_projects(request.user).select_related("firma")
             .annotate(gcount=Count("grafik_rows")).order_by("code"))
    if firma_id:
        projs = projs.filter(firma_id=firma_id)
    rows = [{
        "obj": p,
        "firma": p.firma.name if p.firma_id else "—",
        "gcount": p.gcount,
        "start": p.grafik_start,
    } for p in projs]
    return render(request, "projects/grafik_list.html", {
        "rows": rows, "soni": len(rows),
        "firmalar": visible_firmas(request.user).order_by("name"), "sel_firma": firma_id,
    })


@login_required
def grafik_web(request, pk):
    """Ish grafigi — brauzerda (namunadagi «график работа» kabi tahrirlanadigan jadval)."""
    p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))
    can_edit = is_pto(request.user)
    from django.utils import timezone as _tz
    bugun = _tz.localdate()

    if request.method == "POST":
        if not can_edit:
            raise PermissionDenied("Grafikni faqat PTO tahrirlaydi.")
        # Jadval WYSIWYG: sarlavha sanalari tanlangan boshlanishga qarab jonli
        # yangilanadi — foydalanuvchi qaysi ustunda nimani ko'rsa, shu saqlanadi
        eski_qatorlar = list(p.grafik_rows.all())
        gs = (request.POST.get("grafik_start") or "").strip()
        ge = (request.POST.get("grafik_end") or "").strip()
        try:
            p.grafik_start = datetime.date.fromisoformat(gs) if gs else None
        except ValueError:
            p.grafik_start = None
        try:
            p.grafik_end = datetime.date.fromisoformat(ge) if ge else None
        except ValueError:
            p.grafik_end = None
        p.save(update_fields=["grafik_start", "grafik_end"])
        yangi_start = p.grafik_start or p.start_date or bugun
        N = _grafik_kunlar(yangi_start, bugun, p.grafik_end)

        names = request.POST.getlist("name")
        units = request.POST.getlist("unit")
        qtys = request.POST.getlist("qty")
        plans = request.POST.getlist("plan")
        m_boshlar = request.POST.getlist("muddat_boshi")
        muddatlar = request.POST.getlist("muddat")
        srok_offs = request.POST.getlist("srok_off")
        srok_ons = request.POST.getlist("srok_on")
        resps = request.POST.getlist("responsible")
        notes = request.POST.getlist("note")
        # Forma nechta kun ustunini yuborgan bo'lsa — shunchasini o'qiymiz
        form_n = 0
        while f"d{form_n}" in request.POST and form_n < 400:
            form_n += 1
        daycols = [request.POST.getlist(f"d{j}") for j in range(form_n)]

        CAP = 84
        new_rows = []
        t = 0
        for i in range(len(names)):
            nm = (names[i] or "").strip()
            days = {}
            for j in range(form_n):
                col = daycols[j]
                v = (col[i] if i < len(col) else "").strip()
                if v:
                    d = _to_dec(v)
                    # 0 ham SAQLANADI — «ish bo'lmagan kun» (qizil) shu bilan belgilanadi
                    if d is not None and 0 <= j < CAP:
                        days[str(j)] = float(d)
            # Formada KO'RINMAGAN (oyna qisqarganda tashqarida qolgan) eski kunlar
            # YO'QOLMASIN: o'sha o'rindagi eski qator nomi mos kelsa — saqlab qolamiz
            if i < len(eski_qatorlar) and (eski_qatorlar[i].name or "").strip() == nm:
                for k, v in (eski_qatorlar[i].days or {}).items():
                    try:
                        ki = int(k)
                    except (TypeError, ValueError):
                        continue
                    if ki >= form_n and 0 <= ki < CAP and str(ki) not in days:
                        days[str(ki)] = v
            qv = _to_dec(qtys[i] if i < len(qtys) else "0") or Decimal("0")
            pv = _to_dec(plans[i] if i < len(plans) else "0") or Decimal("0")
            if not nm and not days and qv == 0:
                continue
            mb = None
            try:
                mb_s = (m_boshlar[i] if i < len(m_boshlar) else "").strip()
                mb = datetime.date.fromisoformat(mb_s) if mb_s else None
            except ValueError:
                mb = None
            try:
                md = int(muddatlar[i]) if i < len(muddatlar) and muddatlar[i].strip() else 0
            except ValueError:
                md = 0
            def _idx_list(qatorlar_ro):
                nat = []
                if i < len(qatorlar_ro):
                    for tok in (qatorlar_ro[i] or "").split(","):
                        tok = tok.strip()
                        if tok.isdigit() and 0 <= int(tok) < CAP:
                            nat.append(int(tok))
                return nat
            off = _idx_list(srok_offs)
            qoshimcha = _idx_list(srok_ons)
            t += 1
            new_rows.append(GrafikRow(
                project=p, tartib=t, name=nm[:500],
                unit=(units[i] if i < len(units) else "").strip()[:32],
                qty=qv, plan=pv,
                muddat_boshi=mb, muddat=max(0, min(md, 366)),
                srok_off=off, srok_on=qoshimcha,
                responsible=(resps[i] if i < len(resps) else "").strip()[:255],
                note=(notes[i] if i < len(notes) else "").strip()[:500],
                days=days,
            ))
        with transaction.atomic():
            p.grafik_rows.all().delete()
            GrafikRow.objects.bulk_create(new_rows)
        messages.success(request, f"Grafik saqlandi ({len(new_rows)} qator).")
        return redirect("grafik_web", pk=pk)

    # GET — sana doim ko'rinsin (tanlanmagan bo'lsa: obyekt boshi yoki bugun)
    eff_start = p.grafik_start or p.start_date or bugun
    N = _grafik_kunlar(eff_start, bugun, p.grafik_end)
    # Har bir kun: sana, o'tganmi (bugundan oldin), qaysi haftaga tegishli
    days = []
    for j in range(N):
        d = eff_start + datetime.timedelta(days=j)
        days.append({
            "i": j, "label": d.strftime("%d.%m"), "iso": d.isoformat(),
            "is_past": d < bugun, "week": (d - eff_start).days // 7,
        })
    # Hafta guruhlari — sarlavha ustuni uchun (colspan)
    week_groups = []
    for w in sorted({dd["week"] for dd in days}):
        wdays = [dd for dd in days if dd["week"] == w]
        week_groups.append({
            "week": w, "span": len(wdays),
            "label": f"{wdays[0]['label']} — {wdays[-1]['label']}",
            "all_past": all(dd["is_past"] for dd in wdays),
        })
    day_headers = [dd["label"] for dd in days]
    eski_kun = sum(1 for dd in days if dd["is_past"])
    grows = []
    def _kun_str(v):
        # «%g» katta sonlarni 1.23e+06 ko'rinishida buzardi — aniq format
        s = f"{float(v):.3f}".rstrip("0").rstrip(".")
        return s or "0"
    for r in p.grafik_rows.all():
        kunlar = []
        for j in range(N):
            v = r.days.get(str(j))
            kunlar.append("" if v is None else _kun_str(v))
        grows.append({"r": r, "days": kunlar})
    return render(request, "projects/grafik_web.html", {
        "p": p, "grows": grows, "N_range": list(range(N)), "day_headers": day_headers,
        "days": days, "week_groups": week_groups, "eski_kun": eski_kun,
        "eski_hafta": sum(1 for wg in week_groups if wg["all_past"]),
        "can_edit": can_edit, "N": N,
        "grafik_start": eff_start.isoformat(),
        "grafik_end": p.grafik_end.isoformat() if p.grafik_end else "",
    })


@login_required
def grafik_rabota(request, pk=None):
    """Ish grafigi — namunadagi AYNAN qolipda (grafik_namuna.xlsx dan)."""
    obj_name = "________________________"
    start = None
    fname = "grafik_rabota.xlsx"
    if pk:
        p = _firma_yoki_403(request, get_object_or_404(Project, pk=pk))
        obj_name = p.name
        start = p.start_date
        fname = "grafik_rabota_%s.xlsx" % p.code

    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from django.conf import settings as _s

    tpl = _s.BASE_DIR / "projects" / "xlsx" / "grafik_namuna.xlsx"
    wb = openpyxl.load_workbook(tpl)

    def clear_rows(ws, first_row):
        for row in ws.iter_rows(min_row=first_row, max_row=ws.max_row):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None

    # ===== 1) grafik =====
    ws = wb["график"]
    ws["A1"] = obj_name
    clear_rows(ws, 4)
    # Vebda kiritilgan grafik qatorlari (bo'lsa) — ranglari bilan chiqadi
    grows = list(p.grafik_rows.all()) if pk else []
    if pk and p.grafik_start:
        start = p.grafik_start
    oxirgi = max(21, 3 + len(grows))
    # No + formulalar: kunlik H..V (8..22), E=5, G=7 (foiz), W=23 (qoldiq)
    for i, r in enumerate(range(4, oxirgi + 1), start=1):
        ws.cell(r, 1).value = i
        ws.cell(r, 7).value = f'=IF($E{r}="","",SUM(H{r}:V{r})/$E{r})'
        ws.cell(r, 23).value = f'=IF($E{r}="","",$E{r}-SUM(H{r}:V{r}))'
    # Surilma oyna: grafik haftalik davom etadi, Excel 15 kunlik panelga
    # O'TGAN HAFTA + JORIY HAFTA (+ keyingi kunlar) chiqadi
    ofs = 0
    if grows and start:
        from django.utils import timezone as _tz
        farq = (_tz.localdate() - start).days
        if farq >= 0 and farq // 7 >= 2:
            ofs = (farq // 7 - 1) * 7
    if start:
        ws.cell(3, 8).value = start + datetime.timedelta(days=ofs)
        for j in range(1, 15):
            ws.cell(3, 8 + j).value = f'=IF($H$3="","",$H$3+{j})'
    # Ma'lumotlar + namunadagi ranglar: 0 -> QIZIL, plandan kam -> SARIQ,
    # plan bajarilgan -> YASHIL (bo'sh kun rangsiz)
    if grows:
        from openpyxl.styles import PatternFill
        F_QIZIL = PatternFill("solid", fgColor="FFFF0000")
        F_SARIQ = PatternFill("solid", fgColor="FFFFFF00")
        F_YASHIL = PatternFill("solid", fgColor="FF00B050")
        for idx, g in enumerate(grows):
            r = 4 + idx
            ws.cell(r, 2).value = g.name or None
            ws.cell(r, 4).value = g.unit or None
            ws.cell(r, 5).value = float(g.qty) if g.qty else None
            ws.cell(r, 6).value = float(g.plan) if g.plan else None
            ws.cell(r, 24).value = g.responsible or None
            ws.cell(r, 25).value = g.note or None
            # Foiz va qoldiq — BARCHA haftalar bo'yicha (oyna 15 kun bo'lsa ham)
            ws.cell(r, 7).value = (float(g.bajarilgan) / float(g.qty)) if g.qty else None
            ws.cell(r, 23).value = float(g.qoldiq) if g.qty else None
            plan = float(g.plan or 0)
            for j in range(GRAFIK_N):
                v = (g.days or {}).get(str(ofs + j))
                if v is None:
                    continue
                v = float(v)
                c = ws.cell(r, 8 + j)
                c.value = int(v) if v == int(v) else v
                if v <= 0:
                    c.fill = F_QIZIL
                elif plan > 0 and v < plan:
                    c.fill = F_SARIQ
                else:
                    c.fill = F_YASHIL

    # ===== 2) Solyarka va gaz =====
    if "Солярка ва газ" in wb.sheetnames:
        wf = wb["Солярка ва газ"]
        try:
            wf["A1"] = obj_name
        except Exception:
            pass
        clear_rows(wf, 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="%s"' % fname
    return resp


@login_required
def weekly_export(request, pk):
    """Bitta haftalik so'rovni ALOHIDA Excel qilib yuklab olish."""
    w = get_object_or_404(
        WeeklyRequest.objects.select_related(
            "project", "created_by", "director_by", "approved_by"), pk=pk)
    p = _firma_yoki_403(request, w.project)

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from django.utils.timezone import localtime as _lt

    KIND = dict(KIND_NOMI)
    STATUS = {"draft": "Qoralama", "dir": "Direktor tasdig'ida",
              "submitted": "Admin tasdig'ida", "approved": "Tasdiqlangan"}

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="213145")
    tot_font = Font(bold=True, size=11)
    tot_fill = PatternFill("solid", fgColor="EFF4FF")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '# ##0'
    right = Alignment(horizontal="right")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Haftalik so'rov"
    ws["A1"] = f"{p.code} — {p.name}"
    ws["A1"].font = Font(bold=True, size=14, color="0B1C30")
    davri = f"Haftalik so'rov: {w.week_start:%d.%m.%Y} — {w.week_end:%d.%m.%Y}"
    if w.number:
        davri += f" · {w.number}"
    ws["A2"] = davri + f" · {STATUS.get(w.status, w.status)}"
    ws["A2"].font = Font(size=10, color="64748B")
    imzo = []
    if w.created_by_id:
        imzo.append(f"Kiritdi: {w.created_by.username} · {_lt(w.created_at):%d.%m.%Y %H:%M}")
    if w.director_by_id:
        imzo.append(f"Direktor: {w.director_by.username}"
                    + (f" · {_lt(w.director_at):%d.%m.%Y %H:%M}" if w.director_at else ""))
    if w.approved_by_id:
        imzo.append(f"Admin: {w.approved_by.username}"
                    + (f" · {_lt(w.approved_at):%d.%m.%Y %H:%M}" if w.approved_at else ""))
    ws["A3"] = " | ".join(imzo) if imzo else ""
    ws["A3"].font = Font(size=10, color="64748B")
    ws.append([])
    hrow = 5
    ws.append(["Turi", "Nomi", "Izoh", "Birlik", "Miqdor", "Narxi", "Summa", "Kiritilgan sana"])
    for c in ws[hrow]:
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    d = {k: Decimal("0") for k in KINDS}
    for it in w.items.all():
        d[it.kind] += it.total
        ws.append([
            KIND.get(it.kind, it.kind), it.name, it.note or "", it.unit or "",
            float(it.quantity), float(it.unit_price), float(it.total),
            _lt(it.created_at).strftime("%d.%m.%Y %H:%M") if it.created_at else "",
        ])
    for r in range(hrow + 1, ws.max_row + 1):
        for col in (5, 6, 7):
            ws.cell(r, col).number_format = money
            ws.cell(r, col).alignment = right
        for c in ws[r]:
            c.border = border
    ws.append([])
    for k in KINDS:
        rr = ws.max_row + 1
        ws.cell(rr, 6, KIND[k] + " jami:").font = tot_font
        c = ws.cell(rr, 7, float(d[k]))
        c.number_format = money
        c.font = tot_font
        c.fill = tot_fill
    rr = ws.max_row + 1
    ws.cell(rr, 6, "JAMI:").font = Font(bold=True, size=12)
    c = ws.cell(rr, 7, float(sum(d.values())))
    c.number_format = money
    c.font = Font(bold=True, size=12, color="2563EB")
    c.fill = tot_fill
    ws.freeze_panes = "A6"
    for i, kenglik in enumerate([16, 34, 30, 10, 12, 14, 16, 17], start=1):
        ws.column_dimensions[get_column_letter(i)].width = kenglik

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = ('attachment; filename="haftalik_%s_%s.xlsx"'
                                   % (p.code, w.week_start.strftime("%Y%m%d")))
    return resp


def _find_col(header, *keywords, exclude=()):
    """Sarlavha qatoridan kalit so'z bo'yicha ustun indeksini topadi."""
    for i, cell in enumerate(header):
        t = str(cell).strip().lower() if cell else ""
        if any(k in t for k in keywords) and not any(x in t for x in exclude):
            return i
    return None


@login_required
def limit_import(request):
    """Excel'dan obyekt limitlarini import (Kod bo'yicha yangilash/yaratish)."""
    if not is_pto(request.user):
        raise PermissionDenied("Limit importini faqat PTO bajaradi.")
    natija = None
    if request.method == "POST" and request.FILES.get("file"):
        import openpyxl

        dry = bool(request.POST.get("dry"))
        try:
            wb = openpyxl.load_workbook(request.FILES["file"], data_only=True)
        except Exception:
            return render(request, "projects/limit_import.html", {
                "xato": "Faylni o'qib bo'lmadi — xlsx ekanini tekshiring.",
            })
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return render(request, "projects/limit_import.html", {"xato": "Fayl bo'sh."})

        header = rows[0]
        c_kod = _find_col(header, "kod")
        c_nom = _find_col(header, "nom", "obyekt", "название")
        c_mat = _find_col(header, "material", "материал")
        c_lab = _find_col(header, "ish haqi", "ish", "иш", exclude=("mashina", "машина"))
        c_mach = _find_col(header, "mashina", "chasti", "машина", "механизм")
        c_oth = _find_col(header, "tutilmagan", "kutilmagan", "прочие", "непредвиден")
        if c_kod is None or (c_mat is None and c_lab is None and c_mach is None):
            return render(request, "projects/limit_import.html", {
                "xato": "«Kod» va kamida bitta limit ustuni (Material / Ish haqi / Mashina chasti) topilmadi.",
            })

        def cell(row, idx):
            v = _to_dec(row[idx]) if (idx is not None and idx < len(row)) else None
            return v or Decimal("0")

        admin = is_admin(request.user)
        ozgarishlar = []
        yaratildi = yangilandi = sorov = otkazildi = 0
        for row in rows[1:]:
            if c_kod >= len(row):
                continue
            kod = str(row[c_kod]).strip() if row[c_kod] is not None else ""
            if not kod:
                continue
            mat, lab, mach, oth = cell(row, c_mat), cell(row, c_lab), cell(row, c_mach), cell(row, c_oth)
            jami = mat + lab + mach + oth
            nom = ""
            if c_nom is not None and c_nom < len(row) and row[c_nom]:
                nom = str(row[c_nom]).strip()

            # Firma izolyatsiyasi: PTO faqat o'z firmasidagi obyekt kodini yangilaydi
            p = visible_projects(request.user).filter(code=kod).first()
            if p:
                # Faylda «Ko'zda tutilmagan» ustuni bo'lmasa — mavjud qiymat saqlanadi
                if c_oth is None:
                    oth = p.limit_other or Decimal("0")
                    jami = mat + lab + mach + oth
                eski = (p.limit_material, p.limit_labor, p.limit_machinery, p.limit_other)
                yangi = (mat, lab, mach, oth)
                if eski == yangi:
                    amal = "o'zgarishsiz"
                elif admin:
                    # faqat asosiy admin — to'g'ridan-to'g'ri (PTO birinchi limit ham tasdiq bilan)
                    if not dry:
                        p.limit_material, p.limit_labor, p.limit_machinery, p.limit_other = mat, lab, mach, oth
                        p.save(update_fields=["limit_material", "limit_labor", "limit_machinery", "limit_other"])
                    amal = "yangilandi"
                    yangilandi += 1
                elif p.limit_requests.filter(status__in=LIM_JARAYON).exists():
                    amal = "tasdiq kutmoqda (avvalgi so'rov)"
                else:
                    # PTO mavjud limitni o'zgartiryapti — tasdiqlash zanjiri
                    if not dry:
                        LimitChangeRequest.objects.create(
                            project=p,
                            old_material=p.limit_material, old_labor=p.limit_labor,
                            old_machinery=p.limit_machinery, old_other=p.limit_other,
                            new_material=mat, new_labor=lab, new_machinery=mach, new_other=oth,
                            reason="Excel import", requested_by=request.user,
                            status=_limit_boshlangich(p),
                        )
                    amal = "so'rov yuborildi (tasdiq zanjiri)"
                    sorov += 1
                ozgarishlar.append({"kod": kod, "nomi": p.name, "amal": amal,
                                    "mat": _money(mat), "lab": _money(lab), "mach": _money(mach), "jami": _money(jami)})
            elif admin:
                # yangi obyekt — faqat asosiy admin yaratadi
                if not dry:
                    Project.objects.create(code=kod, name=nom or kod,
                                           limit_material=mat, limit_labor=lab,
                                           limit_machinery=mach, limit_other=oth)
                yaratildi += 1
                ozgarishlar.append({"kod": kod, "nomi": nom or kod, "amal": "yaratildi",
                                    "mat": _money(mat), "lab": _money(lab), "mach": _money(mach), "jami": _money(jami)})
            else:
                # PTO yangi obyekt yarata olmaydi
                otkazildi += 1
                ozgarishlar.append({"kod": kod, "nomi": nom or kod, "amal": "o'tkazib yuborildi (yangi obyekt — admin kerak)",
                                    "mat": _money(mat), "lab": _money(lab), "mach": _money(mach), "jami": _money(jami)})

        natija = {
            "dry": dry,
            "ozgarishlar": ozgarishlar,
            "yaratildi": yaratildi,
            "yangilandi": yangilandi,
            "sorov": sorov,
            "otkazildi": otkazildi,
            "jami": len(ozgarishlar),
        }
    return render(request, "projects/limit_import.html", {"natija": natija})


# ===================== PRORAB -> PTO material so'rovi =====================

def _qty_str(x):
    return _qty(x)


@login_required
def material_sorov(request):
    """Material so'rovlari ro'yxati.
    Prorab — o'zi yuborgan so'rovlar. PTO — o'z obyektlariga kelgan so'rovlar."""
    from .models import MaterialRequest
    vp = visible_projects(request.user)
    reqs = (MaterialRequest.objects.filter(project__in=vp)
            .select_related("project", "created_by", "decided_by")
            .prefetch_related("items"))
    prorab = is_prorab(request.user)
    pto = is_pto(request.user)
    if prorab:
        # Prorab faqat O'ZI yuborgan so'rovlarni ko'radi
        reqs = reqs.filter(created_by=request.user)
    STATUS_CLS = {"pending": "warn", "accepted": "ok", "rejected": "bad"}
    rows = []
    for r in reqs:
        rows.append({
            "obj": r,
            "status_disp": r.get_status_display(),
            "status_cls": STATUS_CLS.get(r.status, ""),
            "soni": r.items.count(),
            "is_pending": r.status == MaterialRequest.Status.PENDING,
        })
    return render(request, "projects/material_sorov.html", {
        "rows": rows,
        "is_prorab": prorab,
        "is_pto": pto,
        "kutayotgan": sum(1 for x in rows if x["is_pending"]),
    })


@login_required
def material_sorov_add(request):
    """Prorab yangi material so'rovini yaratadi va PTOga yuboradi."""
    from .models import MaterialRequest, MaterialRequestItem
    if not (is_prorab(request.user) or request.user.is_superuser):
        raise PermissionDenied("Prorab so'rovini faqat prorab yuboradi.")
    vp = visible_projects(request.user).order_by("code")
    if request.method == "POST":
        pid = request.POST.get("project") or ""
        p = vp.filter(pk=pid).first()
        if not p:
            messages.error(request, "Obyekt tanlang (faqat sizga biriktirilgan).")
            return redirect("material_sorov_add")
        names = request.POST.getlist("name")
        units = request.POST.getlist("unit")
        qtys = request.POST.getlist("quantity")
        notes = request.POST.getlist("note")
        kinds = request.POST.getlist("kind")
        prices = request.POST.getlist("unit_price")
        valid_kind = set(KINDS)
        satrlar = []
        for i in range(len(names)):
            nom = (names[i] or "").strip()
            if not nom:
                continue
            q = _to_dec(qtys[i] if i < len(qtys) else None)
            if q is None or q <= 0:
                continue
            narx = _to_dec(prices[i] if i < len(prices) else None) or Decimal("0")
            if narx < 0:
                narx = Decimal("0")
            satrlar.append(MaterialRequestItem(
                kind=(kinds[i] if i < len(kinds) and kinds[i] in valid_kind else "material"),
                name=nom, unit=(units[i] if i < len(units) else "").strip()[:32],
                quantity=q, unit_price=narx,
                note=(notes[i] if i < len(notes) else "").strip()[:500],
            ))
        if not satrlar:
            messages.error(request, "Kamida bitta material qatori kiriting.")
            return redirect("material_sorov_add")
        req = MaterialRequest.objects.create(
            project=p, created_by=request.user,
            note=(request.POST.get("note_umumiy") or "").strip()[:500],
        )
        for it in satrlar:
            it.request = req
        MaterialRequestItem.objects.bulk_create(satrlar)
        messages.success(request, f"Prorab so'rovi PTOga yuborildi ({len(satrlar)} qator).")
        return redirect("material_sorov")
    return render(request, "projects/material_sorov_add.html", {"obyektlar": vp})


@login_required
def material_sorov_korish(request, pk):
    """So'rovni ko'rish. PTO tahrirlab qabul/rad qilishi mumkin."""
    from .models import MaterialRequest
    req = get_object_or_404(
        MaterialRequest.objects.select_related("project", "created_by", "decided_by"), pk=pk)
    _firma_yoki_403(request, req.project)
    pto = is_pto(request.user)
    items = list(req.items.all())
    jami = Decimal("0")
    for it in items:
        it.price_str = _money(it.unit_price)
        it.total_str = _money(it.total)
        jami += it.total
    return render(request, "projects/material_sorov_korish.html", {
        "req": req,
        "items": items,
        "jami_str": _money(jami),
        "status_disp": req.get_status_display(),
        "can_decide": pto and req.status == MaterialRequest.Status.PENDING,
    })


@login_required
def material_sorov_action(request, pk):
    """PTO material so'rovini tahrirlab QABUL yoki RAD qiladi."""
    from django.utils import timezone

    from .models import MaterialRequest
    req = get_object_or_404(MaterialRequest.objects.select_related("project"), pk=pk)
    _firma_yoki_403(request, req.project)
    if not (is_pto(request.user) or request.user.is_superuser):
        raise PermissionDenied("So'rovni faqat PTO qabul/rad qiladi.")
    if request.method != "POST":
        return redirect("material_sorov_korish", pk=pk)
    action = request.POST.get("action")
    S = MaterialRequest.Status
    if req.status != S.PENDING:
        messages.error(request, "Bu so'rov allaqachon hal qilingan.")
        return redirect("material_sorov_korish", pk=pk)

    # PTO tahriri: qatorlarni yangilash (qabul qilishdan oldin)
    if action in ("accept", "save"):
        valid_kind = set(KINDS)
        for it in req.items.all():
            q = _to_dec(request.POST.get(f"qty_{it.id}"))
            nm = (request.POST.get(f"name_{it.id}") or "").strip()
            un = (request.POST.get(f"unit_{it.id}") or "").strip()[:32]
            kd = (request.POST.get(f"kind_{it.id}") or "").strip()
            pr = _to_dec(request.POST.get(f"price_{it.id}"))
            changed = False
            if q is not None and q > 0 and q != it.quantity:
                it.quantity = q; changed = True
            if nm and nm != it.name:
                it.name = nm; changed = True
            if un != it.unit:
                it.unit = un; changed = True
            if kd in valid_kind and kd != it.kind:
                it.kind = kd; changed = True
            if pr is not None and pr >= 0 and pr != it.unit_price:
                it.unit_price = pr; changed = True
            if changed:
                it.save(update_fields=["quantity", "name", "unit", "kind", "unit_price"])

    if action == "accept":
        req.status = S.ACCEPTED
        req.decided_by = request.user
        req.decided_at = timezone.now()
        req.save(update_fields=["status", "decided_by", "decided_at"])
        messages.success(request, "So'rov qabul qilindi (ro'yxatga saqlandi).")
    elif action == "reject":
        req.status = S.REJECTED
        req.decided_by = request.user
        req.decided_at = timezone.now()
        req.reject_note = (request.POST.get("reject_note") or "").strip()[:500]
        req.save(update_fields=["status", "decided_by", "decided_at", "reject_note"])
        messages.info(request, "So'rov rad etildi.")
    elif action == "save":
        messages.success(request, "O'zgarishlar saqlandi.")
    return redirect("material_sorov_korish", pk=pk)
